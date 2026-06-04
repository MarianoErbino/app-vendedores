"""Build Mapa_Argentina_Shimano_Zonas.html with province/dept polygons, zone fills and 3-level cluster bubbles."""
import os, re, json, sys, unicodedata
from collections import defaultdict
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

MAPS_DIR = r'C:\Users\shimano.sandbox\Desktop\MASTERFILES\PROSPECTOS\MAPAS'
ZONAS_XLSX = r'C:\Users\shimano.sandbox\Desktop\MASTERFILES\ZONAS\TARGETS VENDEDORES-ZONAS.xlsx'
MELI_XLSX = r'C:\Users\shimano.sandbox\Desktop\FORECAST\DATOS_CRUDOS\Masterfile Shimano Venta Ult 365 Días.xlsx'
PRODUCTS_XLSX = r'C:\Users\shimano.sandbox\Desktop\MASTERFILES\PRODUCTO\MASTERFILE PRODUCTOS PESCA.xlsx'
OUT = os.path.join(MAPS_DIR, 'Mapa_Argentina_Shimano_Zonas.html')

# Source provincial files (skip nation and our own previous output)
SKIP = {'Mapa_Argentina_Shimano.html', 'Mapa_Argentina_Shimano_Zonas.html'}

# Province name (in CLIENTES_ZONAS) per provincial-map filename base
prov_from_file = {
    'AMBA': 'BUENOS AIRES',
    'CABA': 'CAPITAL FEDERAL',
    'Catamarca': 'CATAMARCA',
    'Chaco': 'CHACO',
    'Chubut': 'CHUBUT',
    'Cordoba': 'CORDOBA',
    'Corrientes': 'CORRIENTES',
    'EntreRios': 'ENTRE RIOS',
    'Formosa': 'FORMOSA',
    'Jujuy': 'JUJUY',
    'LaPampa': 'LA PAMPA',
    'LaRioja': 'LA RIOJA',
    'Mendoza': 'MENDOZA',
    'Misiones': 'MISIONES',
    'Neuquen': 'NEUQUEN',
    'PBA_Interior': 'BUENOS AIRES',
    'RioNegro': 'RIO NEGRO',
    'Salta': 'SALTA',
    'SanJuan': 'SAN JUAN',
    'SanLuis': 'SAN LUIS',
    'SantaCruz': 'SANTA CRUZ',
    'SantaFe': 'SANTA FE',
    'SantiagoDelEstero': 'SANTIAGO DEL ESTERO',
    'TierraDelFuego': 'TIERRA DEL FUEGO',
    'Tucuman': 'TUCUMAN',
}

# province name used in Mapa_Argentina FeatureCollection -> CLIENTES_ZONAS province
prov_geo_to_zonas = {
    'Buenos Aires': 'BUENOS AIRES',
    'Ciudad Autonoma De Buenos Aires': 'CAPITAL FEDERAL',
    'Catamarca': 'CATAMARCA',
    'Chaco': 'CHACO',
    'Chubut': 'CHUBUT',
    'Cordoba': 'CORDOBA',
    'Corrientes': 'CORRIENTES',
    'Entre Rios': 'ENTRE RIOS',
    'Formosa': 'FORMOSA',
    'Jujuy': 'JUJUY',
    'La Pampa': 'LA PAMPA',
    'La Rioja': 'LA RIOJA',
    'Mendoza': 'MENDOZA',
    'Misiones': 'MISIONES',
    'Neuquen': 'NEUQUEN',
    'Rio Negro': 'RIO NEGRO',
    'Salta': 'SALTA',
    'San Juan': 'SAN JUAN',
    'San Luis': 'SAN LUIS',
    'Santa Cruz': 'SANTA CRUZ',
    'Santa Fe': 'SANTA FE',
    'Santiago Del Estero': 'SANTIAGO DEL ESTERO',
    'Tierra Del Fuego': 'TIERRA DEL FUEGO',
    'Tucuman': 'TUCUMAN',
}

def strip_accents(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def norm(s):
    if s is None:
        return ''
    return strip_accents(str(s)).strip().upper()

def find_var_object(txt, varname):
    pat = re.search(rf'(?:const|var|let)\s+{re.escape(varname)}\s*=\s*', txt)
    if not pat:
        return None
    start = txt.find('{', pat.end())
    if start == -1:
        return None
    depth = 0; i = start; in_str = False; qc = None; esc = False
    while i < len(txt):
        ch = txt[i]
        if in_str:
            if esc: esc = False
            elif ch == chr(92): esc = True
            elif ch == qc: in_str = False; qc = None
        else:
            if ch == '"' or ch == "'": in_str = True; qc = ch
            elif ch == '{': depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0:
                    return txt[start:i+1]
        i += 1
    return None

def js_to_json(s):
    s = re.sub(r'([{,]\s*)([A-Za-z_][A-Za-z_0-9]*)\s*:', r'\1"\2":', s)
    s = re.sub(r',(\s*[}\]])', r'\1', s)
    return s

# === Step 1: parse provincial maps for locationData and departments ===
all_localities = {}   # (province, locality_norm) -> dict
prov_dept_polys = {}  # (province, dept_norm) -> {name, province, geometry}

for fname in sorted(os.listdir(MAPS_DIR)):
    if not fname.endswith('.html') or fname in SKIP:
        continue
    base = fname.replace('Mapa_', '').replace('_Shimano.html', '')
    province = prov_from_file.get(base)
    if province is None:
        continue
    with open(os.path.join(MAPS_DIR, fname), encoding='utf-8') as f:
        txt = f.read()

    loc_js = find_var_object(txt, 'locationData')
    if loc_js:
        try:
            data = json.loads(js_to_json(loc_js))
        except Exception as e:
            print(f'locationData parse error {fname}: {e}')
            data = {}
        for loc_name, loc_data in data.items():
            key = (province, norm(loc_name))
            entry = {
                'province': province,
                'display_name': loc_name,
                'lat': loc_data.get('lat'),
                'lon': loc_data.get('lon'),
                'dept': loc_data.get('dept') or '',
                'clients': list(loc_data.get('clients', [])),
                'prospects': list(loc_data.get('prospects', [])),
            }
            if key in all_localities:
                existing = all_localities[key]
                existing['clients'] = sorted(set(existing['clients'] + entry['clients']))
                existing['prospects'] = sorted(set(existing['prospects'] + entry['prospects']))
            else:
                all_localities[key] = entry

    dep_js = find_var_object(txt, 'departments')
    if dep_js:
        try:
            geo = json.loads(dep_js)
        except Exception as e:
            print(f'departments parse error {fname}: {e}')
            geo = None
        if geo:
            for feat in geo.get('features', []):
                props = feat.get('properties', {})
                dname = props.get('name', '')
                dkey = (province, norm(dname))
                if dkey in prov_dept_polys:
                    continue
                prov_dept_polys[dkey] = {
                    'name': dname,
                    'province': province,
                    'geometry': feat.get('geometry'),
                    'props_extra': {k: v for k, v in props.items() if k != 'name'},
                }

print(f'Localities: {len(all_localities)}  Depts: {len(prov_dept_polys)}')

# === Step 2: province GeoJSON from Mapa_Argentina ===
path_arg = os.path.join(MAPS_DIR, 'Mapa_Argentina_Shimano.html')
with open(path_arg, encoding='utf-8') as f:
    txt_arg = f.read()
prov_geo = json.loads(find_var_object(txt_arg, 'departments'))
# rename provincias to CLIENTES_ZONAS convention
for feat in prov_geo['features']:
    geo_name = feat['properties']['name']
    feat['properties']['name_norm'] = prov_geo_to_zonas.get(geo_name, geo_name.upper())

# === Step 3: parse CLIENTES_ZONAS to assign vendor ===
wb = openpyxl.load_workbook(ZONAS_XLSX, data_only=True)
ws = wb['CLIENTES_ZONAS']
loc_vend_counts = defaultdict(lambda: defaultdict(int))
prov_vend_counts = defaultdict(lambda: defaultdict(int))
client_vend = {}

for r in range(2, ws.max_row + 1):
    prov = norm(ws.cell(r, 4).value)
    vend = norm(ws.cell(r, 5).value)
    loc = norm(ws.cell(r, 8).value)
    name_com = norm(ws.cell(r, 10).value)
    name_fan = norm(ws.cell(r, 11).value)
    if not vend: continue
    if prov in ('CIUDAD AUTONOMA DE BUENOS AIRES', 'CABA'):
        prov = 'CAPITAL FEDERAL'
    prov_vend_counts[prov][vend] += 1
    if loc:
        loc_vend_counts[(prov, loc)][vend] += 1
    if name_com:
        client_vend[name_com] = vend
    if name_fan:
        client_vend[name_fan] = vend

def vendor_for_locality(province, locality_norm, clients_in_loc):
    if (province, locality_norm) in loc_vend_counts:
        return max(loc_vend_counts[(province, locality_norm)].items(), key=lambda x: x[1])[0]
    counts = defaultdict(int)
    for c in clients_in_loc:
        cn = norm(c)
        if cn in client_vend:
            counts[client_vend[cn]] += 1
    if counts:
        return max(counts.items(), key=lambda x: x[1])[0]
    if province in prov_vend_counts and prov_vend_counts[province]:
        return max(prov_vend_counts[province].items(), key=lambda x: x[1])[0]
    return None

# Assign vendor to each locality
for key, loc in all_localities.items():
    loc['vendor'] = vendor_for_locality(loc['province'], key[1], loc['clients'])

# Assign vendor to each dept by majority of locality vendors in that dept
dept_vend_counts = defaultdict(lambda: defaultdict(int))
for loc in all_localities.values():
    if not loc['vendor']:
        continue
    dkey = (loc['province'], norm(loc['dept']))
    dept_vend_counts[dkey][loc['vendor']] += len(loc['clients']) + len(loc['prospects'])

for dkey, dept in prov_dept_polys.items():
    counts = dept_vend_counts.get(dkey, {})
    if counts:
        dept['vendor'] = max(counts.items(), key=lambda x: x[1])[0]
    else:
        # fallback to province majority
        prov = dkey[0]
        if prov in prov_vend_counts and prov_vend_counts[prov]:
            dept['vendor'] = max(prov_vend_counts[prov].items(), key=lambda x: x[1])[0]
        else:
            dept['vendor'] = None

# === Step 4: vendors metadata & colors ===
VENDORS = [
    {'key': 'GONZALO DE LA ROSA',    'zone': 'Z1', 'label': 'Z1 - Gonzalo De La Rosa (CABA + AMBA Norte/Oeste)', 'color': '#00A9E0', 'fill': 'rgba(0,169,224,0.18)'},
    {'key': 'FEDERICO CASTELANELLI', 'zone': 'Z2', 'label': 'Z2 - Federico Castelanelli (AMBA Sur + BA Interior + Costa)', 'color': '#003366', 'fill': 'rgba(0,51,102,0.20)'},
    {'key': 'MARTIN BOIERO',         'zone': 'Z4', 'label': 'Z4 - Martin Boiero (Cordoba + Cuyo + SF Oeste)', 'color': '#E83A2E', 'fill': 'rgba(232,58,46,0.18)'},
    {'key': 'MAURICIO GIL',          'zone': 'Z5', 'label': 'Z5 - Mauricio Gil (Litoral + Norte BA)', 'color': '#F97316', 'fill': 'rgba(249,115,22,0.18)'},
    {'key': 'IOANNIS PALKOUDAKIS',   'zone': 'Z6', 'label': 'Z6 - Ioannis Palkoudakis (Patagonia)', 'color': '#8E44AD', 'fill': 'rgba(142,68,173,0.18)'},
    {'key': 'SANTIAGO ESTEBAN',      'zone': 'Z7', 'label': 'Z7 - Santiago Esteban (NOA + NEA)', 'color': '#F39C12', 'fill': 'rgba(243,156,18,0.20)'},
]
vendor_map = {v['key']: v for v in VENDORS}

# === Step 5: assemble JSON payloads ===
# Points (localities)
points = []
for loc in all_localities.values():
    if not loc.get('lat') or not loc.get('lon'):
        continue
    total = len(loc['clients']) + len(loc['prospects'])
    if total == 0:
        continue
    points.append({
        'name': loc['display_name'],
        'province': loc['province'],
        'lat': loc['lat'],
        'lon': loc['lon'],
        'dept': loc.get('dept', ''),
        'clients': loc['clients'],
        'prospects': loc['prospects'],
        'vendor': loc['vendor'] or '',
    })

# Dept polygons
dept_features = []
for dkey, dept in prov_dept_polys.items():
    if not dept.get('geometry'):
        continue
    dept_features.append({
        'type': 'Feature',
        'geometry': dept['geometry'],
        'properties': {
            'name': dept['name'],
            'province': dept['province'],
            'vendor': dept.get('vendor') or '',
        },
    })
dept_geojson = {'type': 'FeatureCollection', 'features': dept_features}

# Province polygons
prov_features = []
for feat in prov_geo['features']:
    prov_features.append({
        'type': 'Feature',
        'geometry': feat['geometry'],
        'properties': {
            'name': feat['properties']['name'],
            'name_norm': feat['properties']['name_norm'],
        },
    })
prov_geojson = {'type': 'FeatureCollection', 'features': prov_features}

# Centroids
def centroid_of(pts):
    if not pts: return None
    lat = sum(p[0] for p in pts) / len(pts)
    lon = sum(p[1] for p in pts) / len(pts)
    return [lat, lon]

# Vendor centroid + aggregates
vendor_agg = {}
for v in VENDORS:
    vpoints = [p for p in points if p['vendor'] == v['key']]
    coords = [(p['lat'], p['lon']) for p in vpoints]
    c = centroid_of(coords)
    if c is None: continue
    clients_total = sum(len(p['clients']) for p in vpoints)
    prospects_total = sum(len(p['prospects']) for p in vpoints)
    vendor_agg[v['key']] = {
        'lat': c[0], 'lon': c[1],
        'clients': clients_total,
        'prospects': prospects_total,
        'localities': len(vpoints),
    }

# Province centroid per (vendor, province)
prov_agg = []
groups = defaultdict(list)
for p in points:
    groups[(p['vendor'], p['province'])].append(p)
for (vendor, province), pts in groups.items():
    coords = [(p['lat'], p['lon']) for p in pts]
    c = centroid_of(coords)
    prov_agg.append({
        'province': province,
        'vendor': vendor,
        'lat': c[0], 'lon': c[1],
        'clients': sum(len(p['clients']) for p in pts),
        'prospects': sum(len(p['prospects']) for p in pts),
        'localities': len(pts),
    })

print(f'Points: {len(points)}  Provs/vendor: {len(prov_agg)}  Vendors: {len(vendor_agg)}')

# Print per-vendor stats
for v in VENDORS:
    a = vendor_agg.get(v['key'])
    if a:
        print(f"  {v['zone']} {v['key']}: loc={a['localities']} c={a['clients']} p={a['prospects']}")

# === Step 5b: MercadoLibre sales data ===
print('\nLoading MELI sales...')
meli_prov_to_zonas = {
    'Buenos Aires': 'BUENOS AIRES',
    'CABA': 'CAPITAL FEDERAL',
    'Ciudad Autonoma De Buenos Aires': 'CAPITAL FEDERAL',
    'Catamarca': 'CATAMARCA',
    'Chaco': 'CHACO',
    'Chubut': 'CHUBUT',
    'Cordoba': 'CORDOBA',
    'Corrientes': 'CORRIENTES',
    'Entre Rios': 'ENTRE RIOS',
    'Entre Ríos': 'ENTRE RIOS',
    'Formosa': 'FORMOSA',
    'Jujuy': 'JUJUY',
    'La Pampa': 'LA PAMPA',
    'La Rioja': 'LA RIOJA',
    'Mendoza': 'MENDOZA',
    'Misiones': 'MISIONES',
    'Neuquen': 'NEUQUEN',
    'Neuquén': 'NEUQUEN',
    'Rio Negro': 'RIO NEGRO',
    'Río Negro': 'RIO NEGRO',
    'Salta': 'SALTA',
    'San Juan': 'SAN JUAN',
    'San Luis': 'SAN LUIS',
    'Santa Cruz': 'SANTA CRUZ',
    'Santa Fe': 'SANTA FE',
    'Santiago Del Estero': 'SANTIAGO DEL ESTERO',
    'Tierra Del Fuego': 'TIERRA DEL FUEGO',
    'Tucuman': 'TUCUMAN',
}

def name_key(s):
    """Normalize name for matching: alphanumeric uppercase, no spaces/underscores."""
    if s is None: return ''
    s = strip_accents(str(s)).upper()
    return re.sub(r'[^A-Z0-9]', '', s)

# Aggregate MELI: nickname -> {prov: norm_prov, products: {title: qty}}
nick_data = defaultdict(lambda: {'prov': None, 'products': defaultdict(float), 'orders': 0})
try:
    wb_meli = openpyxl.load_workbook(MELI_XLSX, data_only=True, read_only=True)
    ws_meli = wb_meli['Extracto 1']
    rows_seen = 0
    for r_idx, row in enumerate(ws_meli.iter_rows(min_row=2, values_only=True)):
        if not row or row[0] is None: continue
        nickname = row[0]
        seller_prov_raw = row[1]
        title = row[5]
        try: qty = float(row[14]) if row[14] is not None else 0
        except: qty = 0
        seller_prov = meli_prov_to_zonas.get(str(seller_prov_raw).strip(), norm(seller_prov_raw))
        nick_data[nickname]['prov'] = seller_prov
        if title:
            nick_data[nickname]['products'][title] += qty
            nick_data[nickname]['orders'] += 1
        rows_seen += 1
    print(f'  MELI rows: {rows_seen}, unique sellers: {len(nick_data)}')
except Exception as e:
    print(f'  WARN: MELI load failed: {e}')
    nick_data = {}

# Build nickname index for matching: nick_key -> nickname
nick_index = {}
for nick in nick_data.keys():
    nick_index[name_key(nick)] = nick

def match_client_to_meli(client_name):
    nk = name_key(client_name)
    if not nk: return None
    if nk in nick_index:
        return nick_index[nk]
    # try prefix / contains with min length 5
    if len(nk) >= 5:
        for k, nick in nick_index.items():
            if not k: continue
            if k.startswith(nk) or nk.startswith(k):
                if min(len(k), len(nk)) >= 5:
                    return nick
            if (k in nk or nk in k) and min(len(k), len(nk)) >= 6:
                return nick
    return None

# Per-province aggregations: prov -> {title -> {qty, sellers_count}}
prov_products = defaultdict(lambda: defaultdict(lambda: {'qty': 0.0, 'sellers': set()}))
for nick, data in nick_data.items():
    prov = data['prov']
    if not prov: continue
    for title, qty in data['products'].items():
        prov_products[prov][title]['qty'] += qty
        prov_products[prov][title]['sellers'].add(nick)

# Convert to serializable form
prov_top_products = {}
for prov, prods in prov_products.items():
    sorted_prods = sorted(prods.items(), key=lambda x: -x[1]['qty'])[:50]
    prov_top_products[prov] = [
        {'title': t, 'qty': round(d['qty'], 1), 'sellers': len(d['sellers'])}
        for t, d in sorted_prods
    ]

# Match each client/prospect name to MELI nickname; build client_sales dict
client_sales = {}
matched = 0
unique_names = set()
for p in points:
    for n in p['clients'] + p['prospects']:
        unique_names.add((n, p['province']))

for (client_name, prov) in unique_names:
    nick = match_client_to_meli(client_name)
    if nick:
        d = nick_data[nick]
        prods = sorted(d['products'].items(), key=lambda x: -x[1])
        client_sales[client_name + '|' + prov] = {
            'nick': nick,
            'meli_prov': d['prov'],
            'products': [{'title': t, 'qty': round(q, 1)} for t, q in prods],
        }
        matched += 1
print(f'  Matched MELI to {matched} of {len(unique_names)} client names ({matched*100//max(1,len(unique_names))}%)')

# === Step 5c: Load product master ===
print('\nLoading product master...')
products = []
try:
    wb_prod = openpyxl.load_workbook(PRODUCTS_XLSX, data_only=True, read_only=True)
    ws_prod = wb_prod['Hoja1']
    for r in ws_prod.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]: continue
        products.append({
            'code': str(r[0]).strip(),
            'desc': (str(r[1]).strip() if r[1] is not None else ''),
            'cat': (str(r[2]).strip() if r[2] is not None else ''),
            'fam': (str(r[3]).strip() if r[3] is not None else ''),
            'sub': (str(r[4]).strip() if r[4] is not None else ''),
        })
    print(f'  Loaded {len(products)} products')
except Exception as e:
    print(f'  WARN: products load failed: {e}')

# === Step 6: assemble HTML ===
points_json = json.dumps(points, ensure_ascii=False)
client_sales_json = json.dumps(client_sales, ensure_ascii=False)
prov_top_products_json = json.dumps(prov_top_products, ensure_ascii=False)
products_json = json.dumps(products, ensure_ascii=False)
vendors_json = json.dumps([{
    'key': v['key'], 'zone': v['zone'], 'label': v['label'],
    'color': v['color'], 'fill': v['fill'],
} for v in VENDORS], ensure_ascii=False)
vendor_agg_json = json.dumps(vendor_agg, ensure_ascii=False)
prov_agg_json = json.dumps(prov_agg, ensure_ascii=False)
dept_geo_json = json.dumps(dept_geojson, ensure_ascii=False)
prov_geo_json = json.dumps(prov_geojson, ensure_ascii=False)

vendor_options = []
for v in VENDORS:
    vendor_options.append(f'    <option value="{v["key"]}">{v["label"]}</option>')
vendor_options_html = '\n'.join(vendor_options)

html_template = r'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Shimano Fishing - Mapa por Zonas / Vendedores</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:"Segoe UI",Calibri,Arial,sans-serif;background:#f5f7fa;color:#1e293b}
.header{background:#fff;color:#0f172a;padding:14px 28px;display:flex;align-items:center;gap:20px;border-bottom:1px solid #e5e7eb;box-shadow:0 1px 4px rgba(0,0,0,.04);position:relative}
.header .logo-text{font-size:11px;font-weight:800;letter-spacing:3px;color:#fff;background:#00A9E0;padding:6px 14px;border-radius:4px;z-index:1}
.header h1{font-size:20px;font-weight:700;letter-spacing:.5px;position:absolute;left:50%;top:50%;transform:translate(-50%,-50%)}
@media (max-width:900px){.header h1{position:static;transform:none;font-size:15px}}
.controls{background:#fff;padding:10px 18px;display:flex;align-items:center;gap:8px;border-bottom:1px solid #e5e7eb;flex-wrap:nowrap;overflow-x:auto}
.controls label{font-size:11px;font-weight:700;color:#334155;text-transform:uppercase;letter-spacing:.4px;white-space:nowrap}
select.zone-select{padding:6px 10px;border:2px solid #00A9E0;border-radius:6px;background:#fff;color:#0f172a;font-size:12px;font-weight:600;cursor:pointer;min-width:200px;max-width:230px;outline:none}
select.zone-select.small{min-width:120px;max-width:150px;padding:6px 8px;font-size:11px;border-width:1.5px}
select.zone-select:focus{box-shadow:0 0 0 3px rgba(0,169,224,.18)}
.filter-btns{display:flex;gap:4px;margin-left:2px}
.filter-btn{padding:5px 10px;border:1.5px solid #cbd5e1;border-radius:6px;background:#fff;color:#475569;font-size:11px;font-weight:600;cursor:pointer;transition:.15s;white-space:nowrap}
.filter-btn:hover{border-color:#00A9E0;color:#00A9E0}
.filter-btn.active{background:#00A9E0;border-color:#00A9E0;color:#fff}
.btn-export{padding:7px 12px;border:none;border-radius:6px;background:#166534;color:#fff;font-size:11px;font-weight:700;cursor:pointer;text-transform:uppercase;letter-spacing:.4px;display:inline-flex;align-items:center;gap:5px;box-shadow:0 1px 3px rgba(22,101,52,.3);transition:.15s;white-space:nowrap;flex-shrink:0}
.btn-export:hover{background:#14532d;box-shadow:0 2px 6px rgba(22,101,52,.45)}
.btn-export:active{transform:translateY(1px)}
.btn-export::before{content:"\2193";font-size:13px;font-weight:800}
.stats-bar{display:flex;gap:4px;margin-left:auto;margin-right:auto;flex-shrink:0}
.stat-box{text-align:center;padding:3px 10px;background:#f8fafc;border-radius:6px;border:1px solid #e2e8f0}
.stat-box .num{font-size:17px;font-weight:800;color:#00A9E0;line-height:1.1}
.stat-box .lbl{font-size:8px;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.4px}
.main-container{display:flex;height:calc(100vh - 116px)}
#map{flex:1;z-index:1;background:#eaf3f8}
.sidebar{width:420px;background:#fff;overflow-y:auto;border-left:1px solid #e5e7eb;padding:14px 16px;display:flex;flex-direction:column}
.sidebar h3{font-size:13px;color:#0f172a;margin-bottom:10px;padding-bottom:6px;border-bottom:2px solid #00A9E0;font-weight:700;text-transform:uppercase;letter-spacing:.5px}
.tabs{display:flex;gap:4px;margin-bottom:12px}
.tab-btn{flex:1;padding:8px 6px;border:1px solid #e5e7eb;border-bottom:2px solid transparent;background:#f8fafc;color:#64748b;font-size:12px;font-weight:700;cursor:pointer;text-transform:uppercase;letter-spacing:.3px;border-radius:6px 6px 0 0;display:flex;align-items:center;justify-content:center;gap:5px;white-space:nowrap}
.tab-btn:hover{color:#0f172a;background:#f1f5f9}
.tab-btn.active{background:#fff;color:#00A9E0;border-bottom-color:#00A9E0}
.tab-count{background:#e2e8f0;color:#475569;font-size:10px;font-weight:800;padding:1px 7px;border-radius:10px;min-width:22px;text-align:center}
.tab-btn.active .tab-count{background:#00A9E0;color:#fff}
.tab-pane{flex:1;overflow-y:auto}
.contact-controls{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;padding:6px 8px;background:#f8fafc;border-radius:6px;font-size:11px}
.contact-summary{color:#475569;font-weight:600}
.contact-reset{padding:3px 10px;border:1px solid #cbd5e1;border-radius:4px;background:#fff;color:#475569;font-size:10px;font-weight:600;cursor:pointer;text-transform:uppercase;letter-spacing:.5px}
.contact-reset:hover{background:#f1f5f9;color:#0f172a}
.client-card{display:flex;align-items:flex-start;gap:8px;margin-bottom:6px;padding:8px 10px;background:#fff;border:1px solid #e5e7eb;border-left:3px solid #cbd5e1;border-radius:6px;cursor:pointer;transition:.15s}
.client-card:hover{background:#f8fafc}
.client-card.contacted{background:#ecfdf5;border-color:#10b981;border-left-color:#10b981}
.client-card.contacted .client-name{color:#065f46;text-decoration:line-through;text-decoration-thickness:1px}
.client-card input[type=checkbox]{margin-top:2px;width:15px;height:15px;accent-color:#10b981;cursor:pointer;flex-shrink:0}
.client-card .client-body{flex:1;min-width:0}
.client-card .client-name{font-size:12px;font-weight:700;color:#0f172a;word-break:break-word}
.client-card .client-meta{font-size:10px;color:#64748b;margin-top:2px;display:flex;flex-wrap:wrap;gap:6px;align-items:center}
.client-card .badge-tipo{padding:1px 7px;border-radius:8px;font-size:9px;font-weight:800;text-transform:uppercase;letter-spacing:.5px}
.client-card .badge-tipo.cli{background:#dbeafe;color:#1e40af}
.client-card .badge-tipo.pro{background:#fef3c7;color:#92400e}
.client-card.contacted .badge-tipo{opacity:.65}
.loc-card{margin-bottom:8px;padding:10px 12px;background:#fff;border:1px solid #e5e7eb;border-left:3px solid #00A9E0;border-radius:6px;cursor:pointer;transition:.15s}
.loc-card:hover{background:#f0f9fd;border-color:#00A9E0;transform:translateX(2px)}
.loc-card-title{font-size:13px;font-weight:700;color:#0f172a;margin-bottom:3px;display:flex;justify-content:space-between;align-items:center}
.loc-card-title .badge{background:#00A9E0;color:#fff;font-size:10px;font-weight:800;padding:2px 7px;border-radius:10px}
.loc-card-meta{font-size:11px;color:#64748b}
.loc-card-meta .c{color:#0f172a;font-weight:600}
.loc-card-meta .p{color:#64748b;font-weight:600}
.leaflet-popup-content{font-family:"Segoe UI",Arial,sans-serif;font-size:12px}
.leaflet-popup-content-wrapper{border-radius:8px}
.shimano-bubble{cursor:pointer!important}
.leaflet-interactive{cursor:pointer}
.leaflet-marker-icon.shimano-bubble{z-index:9999!important}
.click-debug{position:fixed;top:130px;right:360px;background:#0f172a;color:#fff;padding:8px 14px;border-radius:6px;font-size:12px;font-weight:700;z-index:2000;display:none;box-shadow:0 2px 8px rgba(0,0,0,.35);font-family:monospace}
.modal-overlay{position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(15,23,42,.55);z-index:3000;display:none;align-items:center;justify-content:center;padding:20px}
.modal-overlay.open{display:flex}
.modal-box{background:#fff;border-radius:10px;width:min(900px,96vw);max-height:90vh;overflow:hidden;box-shadow:0 12px 36px rgba(0,0,0,.35);display:flex;flex-direction:column}
.modal-head{padding:16px 22px;border-bottom:1px solid #e5e7eb;display:flex;justify-content:space-between;align-items:flex-start;gap:16px;background:linear-gradient(135deg,#0f172a,#1e293b);color:#fff}
.modal-head h2{font-size:18px;font-weight:800;margin-bottom:4px}
.modal-head .subt{font-size:12px;opacity:.78;display:flex;gap:10px;flex-wrap:wrap}
.modal-head .badge{padding:2px 8px;border-radius:10px;font-size:10px;font-weight:800;letter-spacing:.5px}
.modal-head .badge.tipo{background:rgba(255,255,255,.15)}
.modal-head .badge.prov{background:#00A9E0;color:#fff}
.modal-close{background:transparent;border:none;color:#fff;font-size:24px;cursor:pointer;line-height:1;padding:0 6px}
.modal-close:hover{color:#00A9E0}
.modal-body{padding:18px 22px;overflow-y:auto;display:grid;grid-template-columns:1fr 1fr;gap:20px}
.modal-section{display:flex;flex-direction:column;gap:8px;min-height:0}
.modal-section h3{font-size:12px;font-weight:800;text-transform:uppercase;letter-spacing:.5px;color:#0f172a;padding-bottom:6px;border-bottom:2px solid #00A9E0}
.modal-section.reco h3{border-bottom-color:#10b981;color:#065f46}
.product-row{display:flex;justify-content:space-between;align-items:center;padding:7px 10px;background:#f8fafc;border:1px solid #e5e7eb;border-radius:6px;font-size:11px;gap:8px}
.product-row:hover{background:#f1f5f9}
.product-row .pname{flex:1;color:#1e293b;font-weight:600;word-break:break-word;line-height:1.3}
.product-row .pqty{font-weight:800;color:#00A9E0;font-size:13px;white-space:nowrap}
.product-row.reco{background:#ecfdf5;border-color:#10b981}
.product-row.reco .pqty{color:#10b981}
.product-row .pinfo{font-size:9px;color:#64748b;text-transform:uppercase;letter-spacing:.4px;margin-top:2px}
.modal-footer{padding:10px 22px;border-top:1px solid #e5e7eb;font-size:10px;color:#94a3b8;background:#f8fafc;display:flex;justify-content:space-between}
.no-data{color:#94a3b8;font-size:12px;padding:12px;text-align:center;background:#f8fafc;border-radius:6px}
@media (max-width:700px){.modal-body{grid-template-columns:1fr}}

/* ============ Mobile (<= 900px) ============ */
@media (max-width:900px){
  .header{padding:10px 14px;gap:12px}
  .header h1{font-size:15px;line-height:1.2}
  .header .logo-text{font-size:10px;padding:5px 10px;letter-spacing:2px}
  .controls{padding:8px 12px;gap:6px;flex-wrap:wrap;overflow-x:visible}
  .controls label{font-size:10px}
  select.zone-select{min-width:0;max-width:none;width:100%;font-size:12px}
  select.zone-select.small{min-width:0;max-width:none;width:100%}
  .filter-btns{flex-wrap:wrap;width:100%;justify-content:space-between}
  .filter-btn{flex:1;min-width:0;padding:6px 6px;font-size:11px}
  .stats-bar{margin:0;width:100%;justify-content:space-between;gap:4px;flex-wrap:wrap}
  .stat-box{flex:1;padding:3px 6px}
  .stat-box .num{font-size:15px}
  .stat-box .lbl{font-size:8px}
  .btn-export{width:100%;justify-content:center;padding:9px 12px}
  .main-container{flex-direction:column;height:auto;min-height:calc(100vh - 116px)}
  #map{flex:none;height:55vh;min-height:300px;width:100%}
  .sidebar{width:100%;max-height:none;border-left:none;border-top:1px solid #e5e7eb;padding:12px 14px}
  .tabs{position:sticky;top:0;background:#fff;z-index:5;padding-top:2px}
  .tab-btn{padding:8px 4px;font-size:11px;letter-spacing:.2px}
  .tab-count{font-size:9px;padding:1px 6px}
  /* Modal goes (almost) full screen */
  .modal-overlay{padding:0;align-items:stretch}
  .modal-box{width:100%!important;max-width:none;max-height:100vh;border-radius:0}
  .modal-head{padding:14px 16px}
  .modal-head h2{font-size:16px}
  .modal-head .subt{font-size:11px}
  .modal-body{padding:14px 16px;gap:14px}
  .modal-section h3{font-size:11px}
  /* Pedido modal vertical */
  .pedido-body{grid-template-columns:1fr;overflow-y:auto}
  .pedido-left,.pedido-right{border-right:none;border-bottom:1px solid #e5e7eb;padding:12px 14px}
  .pedido-right{border-bottom:none}
  .filter-row{grid-template-columns:1fr 1fr}
  .product-picker{max-height:35vh}
  .pedido-lines{max-height:none}
  .ped-line{grid-template-columns:65px 1fr 50px 22px;font-size:11px}
  .modal-footer{flex-direction:column;align-items:stretch;gap:8px;padding:10px 14px}
  .modal-footer > div:last-child{display:flex;gap:6px}
  .btn-confirm,.btn-cancel{flex:1;padding:11px 10px;font-size:11px}
  .confirm-box{width:96%;padding:16px 18px}
  .confirm-box h3{font-size:14px}
  /* Suggest box scrolls inside */
  .suggest-box{max-height:200px}
  /* Ocultar burbujas en mobile (ocupan demasiado lugar) */
  .leaflet-marker-pane{display:none}
}

/* ============ Very small phones (<= 480px) ============ */
@media (max-width:480px){
  .header h1{font-size:13px}
  .header .logo-text{font-size:9px;padding:4px 8px}
  .stats-bar{display:grid;grid-template-columns:repeat(4,1fr)}
  .stat-box{padding:3px 4px}
  .filter-row{grid-template-columns:1fr}
  .ped-line{grid-template-columns:56px 1fr 46px 22px}
  .prod-row{grid-template-columns:65px 1fr auto}
  .prod-row .code{font-size:9px}
  .modal-head h2{font-size:14px}
}
.pedido-body{display:grid;grid-template-columns:1.4fr 1fr;gap:0;flex:1;overflow:hidden}
.pedido-left{padding:14px 18px;overflow:hidden;display:flex;flex-direction:column;gap:10px;border-right:1px solid #e5e7eb;background:#f8fafc}
.pedido-right{padding:14px 18px;overflow-y:auto;display:flex;flex-direction:column;gap:8px;background:#fff}
.filter-row{display:grid;grid-template-columns:1fr 1fr;gap:6px}
.filter-row select,.filter-row input{padding:7px 10px;border:1.5px solid #cbd5e1;border-radius:5px;font-size:11px;color:#0f172a;background:#fff;font-family:inherit;outline:none}
.filter-row select:focus,.filter-row input:focus{border-color:#166534;box-shadow:0 0 0 2px rgba(22,101,52,.15)}
.filter-row input{grid-column:1 / -1}
.product-picker{flex:1;overflow-y:auto;background:#fff;border:1px solid #e5e7eb;border-radius:6px;padding:4px}
.prod-row{display:grid;grid-template-columns:80px 1fr auto;gap:8px;align-items:center;padding:6px 8px;font-size:11px;border-bottom:1px solid #f1f5f9;cursor:pointer;transition:.1s}
.prod-row:hover{background:#ecfdf5}
.prod-row:last-child{border-bottom:none}
.prod-row .code{font-family:monospace;font-size:10px;color:#64748b;font-weight:700}
.prod-row .pdesc{color:#0f172a;line-height:1.3;word-break:break-word}
.prod-row .pcat{font-size:9px;color:#94a3b8;text-transform:uppercase;letter-spacing:.4px;display:flex;gap:6px;margin-top:1px}
.prod-row .add-btn{background:#166534;color:#fff;border:none;border-radius:4px;width:24px;height:24px;font-size:14px;font-weight:800;cursor:pointer;line-height:1}
.prod-row .add-btn:hover{background:#15803d}
.prod-row.in-order{background:#ecfdf5;border-left:3px solid #10b981}
.pedido-summary{display:flex;justify-content:space-between;font-size:11px;font-weight:700;color:#475569;padding:8px 10px;background:#f8fafc;border-radius:6px;border:1px solid #e5e7eb}
.pedido-summary span:last-child{color:#166534}
.pedido-lines{display:flex;flex-direction:column;gap:5px;flex:1;overflow-y:auto}
.ped-line{display:grid;grid-template-columns:75px 1fr 56px 24px;gap:6px;align-items:center;padding:6px 8px;background:#f8fafc;border:1px solid #e5e7eb;border-radius:5px;font-size:11px}
.ped-line .pname{color:#0f172a;font-weight:600;line-height:1.25;word-break:break-word}
.ped-line .pcode{font-family:monospace;font-size:10px;color:#64748b;font-weight:700}
.ped-line input.qty{padding:4px 6px;border:1.5px solid #cbd5e1;border-radius:4px;text-align:right;font-size:11px;font-weight:700;color:#0f172a;font-family:inherit;outline:none}
.ped-line input.qty:focus{border-color:#166534}
.ped-line .rm-btn{background:transparent;border:none;color:#dc2626;cursor:pointer;font-size:14px;line-height:1;padding:2px 4px;font-weight:800}
.ped-line .rm-btn:hover{color:#991b1b}
.pedido-client-card{margin-bottom:6px;padding:10px 12px;background:#fff;border:1px solid #e5e7eb;border-left:3px solid #166534;border-radius:6px;cursor:pointer;transition:.15s;display:flex;justify-content:space-between;align-items:center;gap:8px}
.pedido-client-card:hover{background:#f0fdf4;transform:translateX(2px)}
.pedido-client-card .ped-name{font-size:12px;font-weight:700;color:#0f172a;line-height:1.25}
.pedido-client-card .ped-meta{font-size:10px;color:#64748b;margin-top:2px}
.pedido-client-card .ped-badge{background:#166534;color:#fff;font-size:10px;font-weight:800;padding:2px 9px;border-radius:10px;white-space:nowrap}
.pedido-client-card .ped-badge.empty{background:#cbd5e1;color:#475569}
.btn-confirm{padding:9px 18px;border:none;border-radius:6px;background:#166534;color:#fff;font-size:12px;font-weight:800;cursor:pointer;text-transform:uppercase;letter-spacing:.5px;box-shadow:0 1px 3px rgba(22,101,52,.3);transition:.15s}
.btn-confirm:hover:not(:disabled){background:#14532d;box-shadow:0 2px 6px rgba(22,101,52,.45)}
.btn-confirm:disabled{background:#cbd5e1;cursor:not-allowed;box-shadow:none}
.btn-cancel{padding:9px 18px;border:none;border-radius:6px;background:#dc2626;color:#fff;font-size:12px;font-weight:800;cursor:pointer;text-transform:uppercase;letter-spacing:.5px;box-shadow:0 1px 3px rgba(220,38,38,.3);transition:.15s}
.btn-cancel:hover:not(:disabled){background:#991b1b;box-shadow:0 2px 6px rgba(220,38,38,.45)}
.btn-cancel:disabled{background:#cbd5e1;cursor:not-allowed;box-shadow:none}
.confirm-dialog{position:absolute;inset:0;background:rgba(15,23,42,.55);display:none;align-items:center;justify-content:center;padding:20px;z-index:10;border-radius:10px}
.confirm-dialog.open{display:flex}
.confirm-box{background:#fff;border-radius:10px;padding:22px 24px;width:min(440px,94%);box-shadow:0 8px 28px rgba(0,0,0,.3)}
.confirm-box h3{font-size:16px;font-weight:800;color:#0f172a;margin-bottom:4px}
.toggle-pedidos{display:flex;gap:0;margin-bottom:10px;border:1px solid #e5e7eb;border-radius:6px;overflow:hidden}
.toggle-pedidos button{flex:1;padding:7px 8px;border:none;background:#f8fafc;color:#64748b;font-size:11px;font-weight:700;cursor:pointer;text-transform:uppercase;letter-spacing:.4px}
.toggle-pedidos button.active{background:#166534;color:#fff}
.confirmed-card{margin-bottom:6px;padding:10px 12px;background:#fff;border:1px solid #e5e7eb;border-left:3px solid #166534;border-radius:6px;cursor:pointer;transition:.15s}
.confirmed-card:hover{background:#f0fdf4}
.confirmed-card .cc-head{display:flex;justify-content:space-between;align-items:center;gap:8px}
.confirmed-card .cc-name{font-size:12px;font-weight:700;color:#0f172a}
.confirmed-card .cc-month{background:#166534;color:#fff;font-size:10px;font-weight:800;padding:2px 9px;border-radius:10px;text-transform:uppercase;letter-spacing:.4px;white-space:nowrap}
.confirmed-card .cc-meta{font-size:10px;color:#64748b;margin-top:3px}
.confirmed-card .cc-summary{font-size:10px;color:#166534;margin-top:3px;font-weight:600}
.suggest-box{background:linear-gradient(135deg,#ecfdf5,#fef3c7);border:1.5px solid #fbbf24;border-radius:6px;padding:8px 10px;margin-bottom:2px;display:flex;flex-direction:column;gap:6px;max-height:220px;overflow:hidden}
.suggest-box.hidden{display:none}
.suggest-empty-msg{font-size:10px;color:#92400e;background:rgba(255,255,255,.7);padding:6px 8px;border-radius:4px;line-height:1.4}
.suggest-head{display:flex;justify-content:space-between;align-items:center;font-size:10px;font-weight:800;text-transform:uppercase;letter-spacing:.4px;color:#92400e}
.suggest-list{display:flex;flex-direction:column;gap:4px;overflow-y:auto}
.suggest-row{display:grid;grid-template-columns:75px 1fr auto;gap:6px;align-items:center;padding:5px 7px;background:#fff;border:1px solid #fde68a;border-radius:5px;font-size:11px;cursor:pointer;transition:.1s}
.suggest-row:hover{background:#fef3c7;transform:translateX(2px);border-color:#f59e0b}
.suggest-row .code{font-family:monospace;font-size:10px;color:#92400e;font-weight:700}
.suggest-row .sname{color:#0f172a;font-weight:600;line-height:1.25;word-break:break-word}
.suggest-row .sinfo{font-size:9px;color:#92400e;font-weight:600;text-transform:uppercase;letter-spacing:.4px;margin-top:1px}
.suggest-row .add-sg{background:#f59e0b;color:#fff;border:none;border-radius:4px;width:22px;height:22px;font-size:13px;font-weight:800;cursor:pointer;line-height:1}
.suggest-row .add-sg:hover{background:#d97706}
</style>
</head>
<body>
<div class="header">
  <div class="logo-text">SHIMANO</div>
  <h1>Market Scan - Argentina por Zonas</h1>
</div>
<div class="controls">
  <label>Zona:</label>
  <select class="zone-select" id="zone-select">
    <option value="ALL">Todas las zonas</option>
__VENDOR_OPTIONS__
  </select>
  <label>Provincia:</label>
  <select class="zone-select small" id="prov-select"><option value="ALL">Todas</option></select>
  <label>Localidad:</label>
  <select class="zone-select small" id="loc-select"><option value="ALL">Todas</option></select>
  <div class="filter-btns">
    <button class="filter-btn active" data-filter="both" onclick="setFilter('both')">Ambos</button>
    <button class="filter-btn" data-filter="clients" onclick="setFilter('clients')">Clientes</button>
    <button class="filter-btn" data-filter="prospects" onclick="setFilter('prospects')">Prospectos</button>
  </div>
  <div class="stats-bar">
    <div class="stat-box"><div class="num" id="stat-loc">0</div><div class="lbl">Localidades</div></div>
    <div class="stat-box"><div class="num" id="stat-c">0</div><div class="lbl">Clientes</div></div>
    <div class="stat-box"><div class="num" id="stat-p">0</div><div class="lbl">Prospectos</div></div>
    <div class="stat-box"><div class="num" id="stat-t">0</div><div class="lbl">Total</div></div>
  </div>
  <button class="btn-export" onclick="exportToExcel()" title="Descargar Excel con clientes y contactos por vendedor">Exportar a Excel</button>
</div>
<div class="main-container">
  <div id="map"></div>
  <div class="click-debug" id="click-debug"></div>
  <div class="modal-overlay" id="pedido-modal" onclick="if(event.target===this)closePedidoModal()">
    <div class="modal-box" style="width:min(1100px,98vw)">
      <div class="modal-head" style="background:linear-gradient(135deg,#166534,#15803d)">
        <div>
          <h2 id="pm-name">-</h2>
          <div class="subt" id="pm-meta"></div>
        </div>
        <button class="modal-close" onclick="closePedidoModal()" title="Cerrar">&times;</button>
      </div>
      <div class="pedido-body">
        <div class="pedido-left">
          <div class="filter-row">
            <select id="pm-cat" onchange="renderProductPicker()"><option value="ALL">Categoria: Todas</option></select>
            <select id="pm-fam" onchange="renderProductPicker()"><option value="ALL">Familia: Todas</option></select>
            <select id="pm-sub" onchange="renderProductPicker()"><option value="ALL">Subfamilia: Todas</option></select>
            <input type="text" id="pm-search" placeholder="Buscar por nombre o codigo..." oninput="renderProductPicker()" />
          </div>
          <div class="product-picker" id="pm-products"></div>
        </div>
        <div class="pedido-right">
          <div class="pedido-summary"><span id="pm-line-count">0 productos en pedido</span><span id="pm-units">0 unidades</span></div>
          <div class="suggest-box" id="pm-suggest-box">
            <div class="suggest-head"><span>Sugeridos para este cliente</span><span id="pm-suggest-info" style="font-size:9px;color:#94a3b8;font-weight:600"></span></div>
            <div class="suggest-list" id="pm-suggest"></div>
          </div>
          <div style="font-size:10px;font-weight:800;text-transform:uppercase;letter-spacing:.4px;color:#475569;margin-top:4px">Pedido actual</div>
          <div class="pedido-lines" id="pm-lines"></div>
        </div>
      </div>
      <div class="modal-footer" style="justify-content:space-between;gap:12px">
        <div style="display:flex;flex-direction:column;gap:2px"><span>Pedido guardado automaticamente</span><span id="pm-saved-tag" style="color:#475569">Sin cambios</span></div>
        <div style="display:flex;gap:8px">
          <button class="btn-cancel" id="pm-cancel" onclick="cancelPedido()">Cancelar pedido</button>
          <button class="btn-confirm" id="pm-confirm" onclick="openConfirmDialog()">Confirmar pedido</button>
        </div>
      </div>
      <div class="confirm-dialog" id="confirm-dialog">
        <div class="confirm-box">
          <h3 id="cd-title">Confirmar pedido</h3>
          <p id="cd-text" style="font-size:13px;color:#475569;margin:8px 0 14px">Seleccione el mes del pedido:</p>
          <div style="display:flex;gap:8px;margin-bottom:14px">
            <select id="cd-mes" style="flex:2;padding:8px 10px;border:1.5px solid #cbd5e1;border-radius:6px;font-size:13px;font-weight:600;color:#0f172a;outline:none"></select>
            <select id="cd-anio" style="flex:1;padding:8px 10px;border:1.5px solid #cbd5e1;border-radius:6px;font-size:13px;font-weight:600;color:#0f172a;outline:none"></select>
          </div>
          <div style="display:flex;gap:8px;justify-content:flex-end">
            <button onclick="closeConfirmDialog()" style="padding:8px 16px;border:1.5px solid #cbd5e1;border-radius:6px;background:#fff;color:#475569;font-size:12px;font-weight:700;cursor:pointer;text-transform:uppercase;letter-spacing:.4px">Cancelar</button>
            <button onclick="doConfirmPedido()" style="padding:8px 18px;border:none;border-radius:6px;background:#166534;color:#fff;font-size:12px;font-weight:700;cursor:pointer;text-transform:uppercase;letter-spacing:.4px">Confirmar</button>
          </div>
        </div>
      </div>
    </div>
  </div>
  <div class="modal-overlay" id="client-modal" onclick="if(event.target===this)closeClientModal()">
    <div class="modal-box">
      <div class="modal-head">
        <div>
          <h2 id="cm-name">-</h2>
          <div class="subt" id="cm-meta"></div>
        </div>
        <button class="modal-close" onclick="closeClientModal()" title="Cerrar">&times;</button>
      </div>
      <div class="modal-body">
        <div class="modal-section">
          <h3>Historial de compra</h3>
          <div id="cm-history"></div>
        </div>
        <div class="modal-section reco">
          <h3>Seccion recomendados</h3>
          <div id="cm-reco"></div>
        </div>
      </div>
      <div class="modal-footer">
        <span>Fuente: Shimano MELI - ventas ultimos 365 dias</span>
        <span id="cm-source"></span>
      </div>
    </div>
  </div>
  <div class="sidebar">
    <div class="tabs">
      <button class="tab-btn active" data-tab="locs" onclick="setTab('locs')">Localidades <span class="tab-count" id="tab-loc-count">0</span></button>
      <button class="tab-btn" data-tab="clients" onclick="setTab('clients')">Clientes <span class="tab-count" id="tab-cli-count">0</span></button>
      <button class="tab-btn" data-tab="pedidos" onclick="setTab('pedidos')">Pedidos <span class="tab-count" id="tab-ped-count">0</span></button>
    </div>
    <div class="tab-pane" id="pane-locs">
      <h3 id="sidebar-title">Localidades</h3>
      <div id="sidebar-content"></div>
    </div>
    <div class="tab-pane" id="pane-clients" style="display:none">
      <h3 id="clients-title">Clientes en la seleccion</h3>
      <div class="contact-controls">
        <span class="contact-summary" id="contact-summary">0 / 0 contactados</span>
        <button class="contact-reset" onclick="resetContacted()" title="Desmarcar todos">Limpiar</button>
      </div>
      <div id="clients-content"></div>
    </div>
    <div class="tab-pane" id="pane-pedidos" style="display:none">
      <h3>Pedidos</h3>
      <div class="toggle-pedidos">
        <button id="pt-crear" class="active" onclick="setPedidoView('crear')">Crear</button>
        <button id="pt-pend" onclick="setPedidoView('pendientes')">Pendientes</button>
        <button id="pt-conf" onclick="setPedidoView('confirmados')">Confirmados</button>
      </div>
      <div class="contact-controls">
        <span class="contact-summary" id="pedidos-summary">0 clientes</span>
      </div>
      <div id="pedidos-content"></div>
    </div>
  </div>
</div>
<script>
const POINTS = __POINTS_JSON__;
const VENDORS = __VENDORS_JSON__;
const VENDOR_AGG = __VENDOR_AGG_JSON__;
const PROV_AGG = __PROV_AGG_JSON__;
const DEPT_GEO = __DEPT_GEO_JSON__;
const PROV_GEO = __PROV_GEO_JSON__;
const CLIENT_SALES = __CLIENT_SALES_JSON__;
const PROV_TOP_PRODUCTS = __PROV_TOP_PRODUCTS_JSON__;
const PRODUCTS = __PRODUCTS_JSON__;

const vendorLookup = {};
VENDORS.forEach(v => vendorLookup[v.key] = v);

const map = L.map('map', {zoomControl: true, minZoom: 3, maxZoom: 14}).setView([-38, -64], 4);
L.tileLayer('https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png', {
  attribution: '&copy; OSM | CARTO',
  subdomains: 'abcd',
  maxZoom: 18
}).addTo(map);

let currentVendor = 'ALL';
let currentProvince = 'ALL';
let currentLocality = 'ALL';
let currentFilter = 'both';

// === Contacted clients persistence (hoisted para que deptStyle pueda consultar) ===
const CONTACTED_KEY = 'shimano_zonas_contacted_v1';
let contacted = new Set();
try {
  const raw = localStorage.getItem(CONTACTED_KEY);
  if (raw) JSON.parse(raw).forEach(k => contacted.add(k));
} catch(e) { contacted = new Set(); }
function saveContacted(){
  try { localStorage.setItem(CONTACTED_KEY, JSON.stringify([...contacted])); } catch(e) {}
}
function contactKey(p, name, tipo){
  return tipo + '|' + p.province + '|' + p.name + '|' + name;
}

// === Orders persistence (pedidos por cliente contactado) ===
const ORDERS_KEY = 'shimano_zonas_orders_v1';
let orders = {};
try {
  const raw = localStorage.getItem(ORDERS_KEY);
  if (raw) orders = JSON.parse(raw) || {};
} catch(e) { orders = {}; }
function saveOrders(){
  try { localStorage.setItem(ORDERS_KEY, JSON.stringify(orders)); } catch(e) {}
}
function orderKey(clientName, province, locName, tipo){
  return tipo + '|' + province + '|' + locName + '|' + clientName;
}

// === Confirmed orders persistence ===
const CONFIRMED_KEY = 'shimano_zonas_confirmed_v1';
let confirmed = {};
try {
  const raw = localStorage.getItem(CONFIRMED_KEY);
  if (raw) confirmed = JSON.parse(raw) || {};
} catch(e) { confirmed = {}; }
function saveConfirmed(){
  try { localStorage.setItem(CONFIRMED_KEY, JSON.stringify(confirmed)); } catch(e) {}
}
const MESES = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'];

function titleCase(s){
  return String(s).toLowerCase().replace(/\b\w/g, c => c.toUpperCase());
}

function normName(s){
  if (s == null) return '';
  return String(s).normalize('NFD').replace(/[̀-ͯ]/g, '').trim().toUpperCase();
}

// Build dept-name -> matching locality dept normalized
const localityDeptByName = {};
POINTS.forEach(p => {
  const k = p.province + '|' + normName(p.name);
  localityDeptByName[k] = normName(p.dept || '');
});

function deptMatchesLocality(feat){
  if (currentLocality === 'ALL') return true;
  const targetDept = localityDeptByName[currentProvince + '|' + normName(currentLocality)] || '';
  if (!targetDept) return false;
  return normName(feat.properties.name) === targetDept;
}

// Indice (province, normName(dept)) -> array de POINTS dentro de ese departamento
const pointsByDept = {};
POINTS.forEach(p => {
  const k = p.province + '|' + normName(p.dept || '');
  if (!pointsByDept[k]) pointsByDept[k] = [];
  pointsByDept[k].push(p);
});

function deptContactPct(feat){
  // % de clientes+prospectos contactados entre las localidades del dept
  const k = feat.properties.province + '|' + normName(feat.properties.name);
  const pts = pointsByDept[k] || [];
  let total = 0, done = 0;
  pts.forEach(p => {
    p.clients.forEach(n => { total++; if (contacted && contacted.has(contactKey(p, n, 'C'))) done++; });
    p.prospects.forEach(n => { total++; if (contacted && contacted.has(contactKey(p, n, 'P'))) done++; });
  });
  if (total === 0) return -1; // sin clientes en el dept
  return done / total;
}

function deptStyle(feat){
  const v = feat.properties.vendor;
  const vm = vendorLookup[v];
  const GREEN = '#10b981';
  if (currentVendor !== 'ALL') {
    if (v !== currentVendor) {
      return {fillOpacity: 0, opacity: 0, weight: 0, interactive: false};
    }
    if (currentProvince !== 'ALL' && feat.properties.province !== currentProvince) {
      return {fillOpacity: 0, opacity: 0, weight: 0, interactive: false};
    }
    if (!deptMatchesLocality(feat)) {
      return {fillOpacity: 0, opacity: 0, weight: 0, interactive: false};
    }
    const pct = deptContactPct(feat);
    const fully = pct >= 1;
    const fill = fully ? GREEN : (vm ? vm.color : '#cbd5e1');
    return {
      fillColor: fill,
      fillOpacity: fully ? 0.55 : (currentLocality !== 'ALL' ? 0.42 : 0.30),
      color: fully ? GREEN : (vm ? vm.color : '#94a3b8'),
      weight: fully ? 1.4 : (currentLocality !== 'ALL' ? 1.2 : 0.6),
      opacity: 0.9,
    };
  }
  // currentVendor === 'ALL'
  if (currentProvince !== 'ALL' && feat.properties.province !== currentProvince) {
    return {fillOpacity: 0, opacity: 0, weight: 0, interactive: false};
  }
  if (!deptMatchesLocality(feat)) {
    return {fillOpacity: 0, opacity: 0, weight: 0, interactive: false};
  }
  const pct = deptContactPct(feat);
  const fully = pct >= 1;
  const fill = fully ? GREEN : (vm ? vm.color : '#cbd5e1');
  return {
    fillColor: fill,
    fillOpacity: fully ? 0.5 : (vm ? 0.22 : 0.05),
    color: fully ? GREEN : '#94a3b8',
    weight: fully ? 1.0 : 0.4,
    opacity: fully ? 0.9 : 0.6,
  };
}

const deptLayer = L.geoJSON(DEPT_GEO, {
  style: deptStyle,
  interactive: false,
}).addTo(map);

// Build set of provinces (normalized) per vendor for prov visibility
const vendorProvinces = {};
VENDORS.forEach(v => vendorProvinces[v.key] = new Set());
DEPT_GEO.features.forEach(f => {
  const v = f.properties.vendor;
  if (v && vendorProvinces[v]) vendorProvinces[v].add(f.properties.province);
});

function provStyle(feat){
  if (currentVendor !== 'ALL') {
    // Cuando se filtra un vendedor, ocultamos todos los bordes provinciales:
    // solo deben verse los departamentos que pertenecen a su zona.
    return {fillOpacity: 0, opacity: 0, weight: 0};
  }
  return {fillColor: 'transparent', fillOpacity: 0, color: '#475569', weight: 1.2, opacity: 0.85};
}

const provLayer = L.geoJSON(PROV_GEO, {
  style: provStyle,
  interactive: false,
}).addTo(map);

function restyleZoneLayers(){
  deptLayer.setStyle(deptStyle);
  provLayer.setStyle(provStyle);
}

let markerLayer = L.layerGroup().addTo(map);

function makeBubbleIcon(label, color, size, selected, contactPct){
  // contactPct: 0..1 fraccion de clientes contactados en la localidad
  contactPct = contactPct || 0;
  const radius = size / 2;
  const fontSize = String(label).length > 3 ? 11 : String(label).length > 2 ? 12 : 13;
  const border = selected ? 4 : 2.4;
  const green = '#10b981';
  // Si todos contactados: burbuja verde llena. Si parcial: vendor color + dot verde.
  let bg, fg, brd;
  if (contactPct >= 1) {
    bg = green; fg = '#fff'; brd = green;
  } else {
    bg = '#fff'; fg = color; brd = color;
  }
  const ring = selected ? ';box-shadow:0 0 0 4px rgba(0,169,224,.35),0 2px 6px rgba(0,0,0,.25)' : ';box-shadow:0 1px 5px rgba(0,0,0,.22)';
  let html = '<div style="position:relative;background:' + bg + ';color:' + fg + ';border:' + border + 'px solid ' + brd + ';border-radius:50%;width:' + size + 'px;height:' + size + 'px;display:flex;align-items:center;justify-content:center;font-size:' + fontSize + 'px;font-weight:800;font-family:Arial;cursor:pointer;pointer-events:auto' + ring + '">';
  html += label;
  // dot verde si es parcial (no 0, no 100)
  if (contactPct > 0 && contactPct < 1) {
    const dotSize = Math.max(10, size * 0.32);
    html += '<div style="position:absolute;top:-3px;right:-3px;width:' + dotSize + 'px;height:' + dotSize + 'px;background:' + green + ';border:2px solid #fff;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:' + Math.round(dotSize*0.55) + 'px;color:#fff;font-weight:800">&#10003;</div>';
  }
  html += '</div>';
  return L.divIcon({className:'shimano-bubble', html: html, iconSize:[size, size], iconAnchor:[radius, radius]});
}

function localityContactPct(p){
  const all = p.clients.length + p.prospects.length;
  if (all === 0) return 0;
  let done = 0;
  p.clients.forEach(n => { if (contacted.has(contactKey(p, n, 'C'))) done++; });
  p.prospects.forEach(n => { if (contacted.has(contactKey(p, n, 'P'))) done++; });
  return done / all;
}

function getCounts(p){
  return {
    c: currentFilter !== 'prospects' ? p.clients.length : 0,
    pr: currentFilter !== 'clients' ? p.prospects.length : 0,
  };
}

function filteredPoints(){
  return POINTS.filter(p => {
    if (currentVendor !== 'ALL' && p.vendor !== currentVendor) return false;
    if (currentProvince !== 'ALL' && p.province !== currentProvince) return false;
    if (currentLocality !== 'ALL' && p.name !== currentLocality) return false;
    const {c, pr} = getCounts(p);
    return (c + pr) > 0;
  });
}

function populateProvinces(){
  const sel = document.getElementById('prov-select');
  const provs = new Set();
  POINTS.forEach(p => {
    if (currentVendor !== 'ALL' && p.vendor !== currentVendor) return;
    provs.add(p.province);
  });
  const sorted = [...provs].sort();
  sel.innerHTML = '<option value="ALL">Todas</option>' + sorted.map(p => '<option value="' + p + '">' + titleCase(p) + '</option>').join('');
  if (currentProvince !== 'ALL' && !provs.has(currentProvince)) currentProvince = 'ALL';
  sel.value = currentProvince;
}

function populateLocalities(){
  const sel = document.getElementById('loc-select');
  const locs = [];
  POINTS.forEach(p => {
    if (currentVendor !== 'ALL' && p.vendor !== currentVendor) return;
    if (currentProvince !== 'ALL' && p.province !== currentProvince) return;
    locs.push(p.name);
  });
  locs.sort((a, b) => a.localeCompare(b));
  sel.innerHTML = '<option value="ALL">Todas</option>' + locs.map(p => '<option value="' + p.replace(/"/g, '&quot;') + '">' + p + '</option>').join('');
  if (currentLocality !== 'ALL' && !locs.includes(currentLocality)) currentLocality = 'ALL';
  sel.value = currentLocality;
}

function escapeHtml(s){
  return String(s).replace(/[&<>"']/g, ch => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[ch]));
}

function buildLocalityPopup(p){
  const c = p.clients.length, pr = p.prospects.length;
  const vm = vendorLookup[p.vendor];
  const accent = vm ? vm.color : '#00A9E0';
  let html = '<div style="min-width:230px;max-width:300px">';
  html += '<div style="font-size:14px;font-weight:700;color:#0f172a;margin-bottom:2px">' + escapeHtml(p.name) + '</div>';
  html += '<div style="font-size:10px;font-weight:700;color:' + accent + ';text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px">' + escapeHtml(p.province) + (p.dept ? ' - ' + escapeHtml(p.dept) : '') + '</div>';
  if (p.vendor) {
    html += '<div style="font-size:10px;color:#64748b;margin-bottom:6px"><b>Vendedor:</b> ' + escapeHtml(titleCase(p.vendor)) + '</div>';
  }
  html += '<div style="display:flex;gap:10px;margin-bottom:6px">';
  html += '<div style="text-align:center;flex:1;padding:6px;background:#f8fafc;border-radius:4px"><div style="font-size:18px;font-weight:800;color:' + accent + '">' + c + '</div><div style="font-size:8px;color:#64748b">CLIENTES</div></div>';
  html += '<div style="text-align:center;flex:1;padding:6px;background:#f8fafc;border-radius:4px"><div style="font-size:18px;font-weight:800;color:#0f172a">' + pr + '</div><div style="font-size:8px;color:#64748b">PROSPECTOS</div></div>';
  html += '<div style="text-align:center;flex:1;padding:6px;background:#f8fafc;border-radius:4px"><div style="font-size:18px;font-weight:800;color:' + accent + '">' + (c + pr) + '</div><div style="font-size:8px;color:#64748b">TOTAL</div></div>';
  html += '</div>';
  if (currentFilter !== 'prospects' && p.clients.length) {
    html += '<div style="font-size:9px;font-weight:800;color:' + accent + ';text-transform:uppercase;margin-top:6px">Clientes actuales</div>';
    html += '<ul style="margin:2px 0 0;padding:0;list-style:none;max-height:120px;overflow-y:auto">';
    for (const cl of p.clients) html += '<li style="font-size:11px;padding:1px 0;color:#1e293b">- ' + escapeHtml(cl) + '</li>';
    html += '</ul>';
  }
  if (currentFilter !== 'clients' && p.prospects.length) {
    html += '<div style="font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;margin-top:6px">Prospectos</div>';
    html += '<ul style="margin:2px 0 0;padding:0;list-style:none;max-height:120px;overflow-y:auto">';
    for (const pr2 of p.prospects) html += '<li style="font-size:11px;padding:1px 0;color:#475569;font-style:italic">- ' + escapeHtml(pr2) + '</li>';
    html += '</ul>';
  }
  html += '</div>';
  return html;
}

function buildAggPopup(title, subtitle, c, pr, color, extra){
  let html = '<div style="min-width:200px">';
  html += '<div style="font-size:14px;font-weight:700;color:#0f172a;margin-bottom:2px">' + escapeHtml(title) + '</div>';
  if (subtitle) html += '<div style="font-size:10px;font-weight:700;color:' + color + ';text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px">' + escapeHtml(subtitle) + '</div>';
  html += '<div style="display:flex;gap:10px;margin-bottom:6px">';
  html += '<div style="text-align:center;flex:1;padding:6px;background:#f8fafc;border-radius:4px"><div style="font-size:18px;font-weight:800;color:' + color + '">' + c + '</div><div style="font-size:8px;color:#64748b">CLIENTES</div></div>';
  html += '<div style="text-align:center;flex:1;padding:6px;background:#f8fafc;border-radius:4px"><div style="font-size:18px;font-weight:800;color:#0f172a">' + pr + '</div><div style="font-size:8px;color:#64748b">PROSPECTOS</div></div>';
  html += '<div style="text-align:center;flex:1;padding:6px;background:#f8fafc;border-radius:4px"><div style="font-size:18px;font-weight:800;color:' + color + '">' + (c + pr) + '</div><div style="font-size:8px;color:#64748b">TOTAL</div></div>';
  html += '</div>';
  if (extra) html += '<div style="font-size:10px;color:#64748b;margin-top:4px">' + extra + '</div>';
  html += '</div>';
  return html;
}

function getDisplayMode(){
  const z = map.getZoom();
  if (currentVendor === 'ALL') {
    if (z <= 5) return 'vendor';
    if (z <= 7) return 'province';
    return 'locality';
  } else {
    // when vendor selected: skip vendor-level (only show province / locality)
    if (z <= 6) return 'province';
    return 'locality';
  }
}

function drawMarkers(){
  markerLayer.clearLayers();
  const mode = getDisplayMode();
  const visiblePoints = filteredPoints();

  if (mode === 'vendor') {
    const groups = {};
    for (const p of visiblePoints) {
      const k = p.vendor || 'OTRO';
      if (!groups[k]) groups[k] = {pts:[], c:0, pr:0};
      const cc = getCounts(p);
      groups[k].pts.push(p);
      groups[k].c += cc.c;
      groups[k].pr += cc.pr;
    }
    for (const [vkey, g] of Object.entries(groups)) {
      const total = g.c + g.pr;
      if (total === 0) continue;
      const vm = vendorLookup[vkey] || {color:'#00A9E0', zone:'?', label:vkey};
      const lat = g.pts.reduce((s,p)=>s+p.lat,0) / g.pts.length;
      const lon = g.pts.reduce((s,p)=>s+p.lon,0) / g.pts.length;
      const size = Math.max(46, Math.min(72, 30 + Math.sqrt(total) * 3.6));
      const icon = makeBubbleIcon(total, vm.color, size);
      const popup = buildAggPopup(vm.zone + ' - ' + titleCase(vkey), vm.label, g.c, g.pr, vm.color, g.pts.length + ' localidades');
      const m = L.marker([lat, lon], {icon});
      const vk = vkey;
      m.on('click', function(){
        showClickDebug('Click vendedor: ' + vk);
        selectFromMap({vendor: vk});
      });
      m.addTo(markerLayer);
    }
  } else if (mode === 'province') {
    const groups = {};
    for (const p of visiblePoints) {
      const k = (p.vendor || '_') + '|' + p.province;
      if (!groups[k]) groups[k] = {pts:[], c:0, pr:0, vendor:p.vendor, province:p.province};
      const cc = getCounts(p);
      groups[k].pts.push(p);
      groups[k].c += cc.c;
      groups[k].pr += cc.pr;
    }
    for (const g of Object.values(groups)) {
      const total = g.c + g.pr;
      if (total === 0) continue;
      const vm = vendorLookup[g.vendor] || {color:'#00A9E0'};
      const lat = g.pts.reduce((s,p)=>s+p.lat,0) / g.pts.length;
      const lon = g.pts.reduce((s,p)=>s+p.lon,0) / g.pts.length;
      const size = Math.max(36, Math.min(58, 24 + Math.sqrt(total) * 3));
      const icon = makeBubbleIcon(total, vm.color, size);
      const popup = buildAggPopup(titleCase(g.province), 'Vendedor: ' + titleCase(g.vendor || 'Sin asignar'), g.c, g.pr, vm.color, g.pts.length + ' localidades');
      const m = L.marker([lat, lon], {icon});
      const gv = g.vendor, gp = g.province;
      m.on('click', function(){
        showClickDebug('Click provincia: ' + gp);
        selectFromMap({vendor: gv, province: gp});
      });
      m.addTo(markerLayer);
    }
  } else { // locality
    for (const p of visiblePoints) {
      const cc = getCounts(p);
      const total = cc.c + cc.pr;
      if (total === 0) continue;
      const vm = vendorLookup[p.vendor] || {color:'#00A9E0'};
      const isSelected = currentLocality === p.name && (currentProvince === 'ALL' || currentProvince === p.province);
      const baseSize = Math.max(24, Math.min(44, 18 + Math.sqrt(total) * 3));
      const size = isSelected ? baseSize + 6 : baseSize;
      const pct = localityContactPct(p);
      const icon = makeBubbleIcon(total, vm.color, size, isSelected, pct);
      const m = L.marker([p.lat, p.lon], {icon});
      const data = {vendor: p.vendor, province: p.province, name: p.name};
      m.on('click', function(){
        showClickDebug('Click localidad: ' + data.name);
        selectFromMap({vendor: data.vendor, province: data.province, locality: data.name});
      });
      m.addTo(markerLayer);
    }
  }
  updateStats(visiblePoints);
  updateSidebar(visiblePoints);
}

function updateStats(pts){
  let c = 0, p = 0;
  for (const pt of pts) { c += pt.clients.length; p += pt.prospects.length; }
  document.getElementById('stat-loc').textContent = pts.length;
  document.getElementById('stat-c').textContent = currentFilter === 'prospects' ? '-' : c;
  document.getElementById('stat-p').textContent = currentFilter === 'clients' ? '-' : p;
  document.getElementById('stat-t').textContent = (currentFilter === 'clients' ? c : currentFilter === 'prospects' ? p : c + p);
}

function toggleContact(key){
  if (contacted.has(key)) contacted.delete(key); else contacted.add(key);
  saveContacted();
  updateContactSummary();
}
window.toggleContact = toggleContact;

function resetContacted(){
  if (!confirm('Desmarcar todos los clientes contactados de la seleccion actual?')) return;
  const pts = filteredPoints();
  const toRemove = [];
  contacted.forEach(k => {
    const [tipo, prov, locname] = k.split('|');
    if (pts.some(p => p.province === prov && p.name === locname)) toRemove.push(k);
  });
  toRemove.forEach(k => contacted.delete(k));
  saveContacted();
  renderClients(filteredPoints());
  restyleZoneLayers();
  drawMarkers();
}
window.resetContacted = resetContacted;

function updateContactSummary(){
  const pts = filteredPoints();
  let total = 0, done = 0;
  pts.forEach(p => {
    p.clients.forEach(n => { total++; if (contacted.has(contactKey(p, n, 'C'))) done++; });
    p.prospects.forEach(n => { total++; if (contacted.has(contactKey(p, n, 'P'))) done++; });
  });
  const el = document.getElementById('contact-summary');
  if (el) el.textContent = done + ' / ' + total + ' contactados';
  document.getElementById('tab-cli-count').textContent = total;
}

function renderClients(pts){
  const titleEl = document.getElementById('clients-title');
  if (currentLocality !== 'ALL') titleEl.textContent = 'Clientes en ' + currentLocality;
  else if (currentProvince !== 'ALL') titleEl.textContent = 'Clientes en ' + titleCase(currentProvince);
  else if (currentVendor !== 'ALL') {
    const v = vendorLookup[currentVendor];
    titleEl.textContent = 'Clientes - ' + (v ? v.zone : '') + ' ' + titleCase(currentVendor);
  } else titleEl.textContent = 'Clientes (vista nacional)';

  const items = [];
  pts.forEach(p => {
    if (currentFilter !== 'prospects') {
      p.clients.forEach(n => items.push({name:n, tipo:'C', point:p}));
    }
    if (currentFilter !== 'clients') {
      p.prospects.forEach(n => items.push({name:n, tipo:'P', point:p}));
    }
  });
  // sort: contactados primero, luego pendientes; alfabetico dentro de cada grupo
  items.sort((a, b) => {
    const ka = contactKey(a.point, a.name, a.tipo);
    const kb = contactKey(b.point, b.name, b.tipo);
    const da = contacted.has(ka), db = contacted.has(kb);
    if (da !== db) return da ? -1 : 1;
    return a.name.localeCompare(b.name);
  });
  let html = '';
  for (const it of items) {
    const k = contactKey(it.point, it.name, it.tipo);
    const done = contacted.has(k);
    const tipoLbl = it.tipo === 'C' ? 'Cliente' : 'Prospecto';
    const tipoCls = it.tipo === 'C' ? 'cli' : 'pro';
    const vm = vendorLookup[it.point.vendor];
    const accent = vm ? vm.color : '#00A9E0';
    const safeName = escapeAttr(it.name);
    const safeProv = escapeAttr(it.point.province);
    const safeLoc = escapeAttr(it.point.name);
    html += '<div class="client-card' + (done ? ' contacted' : '') + '"';
    html += ' data-key="' + escapeHtml(k) + '"';
    html += ' style="border-left-color:' + (done ? '#10b981' : accent) + '" title="Doble click: ver historial y recomendaciones"';
    html += ' onclick="onClientCardClick(event, \'' + escapeAttr(k) + '\', ' + it.point.lat + ', ' + it.point.lon + ')"';
    html += ' ondblclick="openClientModal(\'' + safeName + '\', \'' + it.tipo + '\', \'' + safeProv + '\', \'' + safeLoc + '\')">';
    html += '<input type="checkbox" ' + (done ? 'checked' : '') + ' onclick="event.stopPropagation();onClientCheck(\'' + escapeAttr(k) + '\', this)"/>';
    html += '<div class="client-body">';
    html += '<div class="client-name">' + escapeHtml(it.name) + '</div>';
    html += '<div class="client-meta">';
    html += '<span class="badge-tipo ' + tipoCls + '">' + tipoLbl + '</span>';
    html += '<span>' + escapeHtml(it.point.name) + '</span>';
    html += '<span style="color:#94a3b8">/ ' + escapeHtml(it.point.province) + '</span>';
    html += '</div></div></div>';
  }
  if (!items.length) html = '<div style="color:#94a3b8;font-size:12px;padding:8px">Sin clientes para esta combinacion.</div>';
  document.getElementById('clients-content').innerHTML = html;
  updateContactSummary();
}

function escapeAttr(s){
  return String(s).replace(/[\\']/g, '\\$&').replace(/"/g, '&quot;');
}

window.onClientCheck = function(key, cb){
  toggleContact(key);
  const card = cb.closest('.client-card');
  if (card) card.classList.toggle('contacted', contacted.has(key));
  // Refrescar burbujas + polígono del departamento + tab pedidos
  if (typeof restyleZoneLayers === 'function') restyleZoneLayers();
  if (typeof drawMarkers === 'function') drawMarkers();
  if (typeof renderPedidosTab === 'function') renderPedidosTab();
};

window.onClientCardClick = function(ev, key, lat, lon){
  if (ev.target.tagName === 'INPUT') return;
  map.flyTo([lat, lon], 12, {duration: .7});
};

function openClientModal(clientName, tipo, province, localityName){
  const tipoLbl = tipo === 'C' ? 'Cliente actual' : 'Prospecto';
  document.getElementById('cm-name').textContent = clientName;
  document.getElementById('cm-meta').innerHTML = '<span class="badge tipo">' + tipoLbl + '</span>'
    + '<span class="badge prov">' + escapeHtml(titleCase(province)) + '</span>'
    + '<span>' + escapeHtml(localityName) + '</span>';

  const salesKey = clientName + '|' + province;
  const own = CLIENT_SALES[salesKey];
  const histEl = document.getElementById('cm-history');
  let ownTitles = new Set();
  if (own && own.products.length) {
    histEl.innerHTML = own.products.map(p => {
      ownTitles.add(p.title);
      return '<div class="product-row"><div class="pname">' + escapeHtml(p.title) + '</div><div class="pqty">' + p.qty + '</div></div>';
    }).join('');
    document.getElementById('cm-source').textContent = 'Matched: ' + own.nick + ' (' + (own.meli_prov || 's/d') + ')';
  } else {
    histEl.innerHTML = '<div class="no-data">Sin historial de venta MELI para este cliente.<br><span style="font-size:10px">(Posiblemente no opera en MercadoLibre o el nombre no matchea con su nickname)</span></div>';
    document.getElementById('cm-source').textContent = '';
  }

  // Recomendaciones: top productos de la provincia que el cliente NO tiene
  const recoEl = document.getElementById('cm-reco');
  const provProds = PROV_TOP_PRODUCTS[province] || [];
  // peers en la provincia (excluyendo a este cliente)
  const peerCount = new Set();
  Object.values(CLIENT_SALES).forEach(s => { if (s.meli_prov === province && (!own || s.nick !== own.nick)) peerCount.add(s.nick); });
  const peers = peerCount.size;
  const reco = [];
  for (const p of provProds) {
    if (ownTitles.has(p.title)) continue;
    reco.push(p);
    if (reco.length >= 12) break;
  }
  if (reco.length) {
    recoEl.innerHTML = '<div style="font-size:10px;color:#475569;margin-bottom:6px">' + peers + ' casa(s) de pesca en ' + titleCase(province) + ' compraron estos productos que <b>' + escapeHtml(clientName) + '</b> NO tiene:</div>'
      + reco.map(p => '<div class="product-row reco"><div><div class="pname">' + escapeHtml(p.title) + '</div><div class="pinfo">' + p.sellers + ' casa(s) lo venden</div></div><div class="pqty">+' + p.qty + '</div></div>').join('');
  } else {
    recoEl.innerHTML = '<div class="no-data">' + (provProds.length ? 'El cliente ya cubre todos los productos top de la provincia.' : 'Sin datos de la provincia para generar recomendaciones.') + '</div>';
  }
  document.getElementById('client-modal').classList.add('open');
}
window.openClientModal = openClientModal;

function closeClientModal(){
  document.getElementById('client-modal').classList.remove('open');
}
window.closeClientModal = closeClientModal;

window.exportToExcel = function(){
  if (typeof XLSX === 'undefined') {
    alert('La libreria de Excel no se cargo. Verifique su conexion a internet y reintente.');
    return;
  }
  const wb = XLSX.utils.book_new();

  // Aggregate per vendor
  const perVendor = {};
  VENDORS.forEach(v => {
    perVendor[v.key] = {
      zone: v.zone, vendor: v.key,
      provincias: new Set(), localidades: new Set(),
      clientes_total: 0, clientes_contactados: 0,
      prospectos_total: 0, prospectos_contactados: 0,
      rows: [],
    };
  });
  perVendor['SIN ASIGNAR'] = {
    zone: '-', vendor: 'SIN ASIGNAR',
    provincias: new Set(), localidades: new Set(),
    clientes_total: 0, clientes_contactados: 0,
    prospectos_total: 0, prospectos_contactados: 0,
    rows: [],
  };

  POINTS.forEach(p => {
    const vk = p.vendor || 'SIN ASIGNAR';
    const agg = perVendor[vk];
    if (!agg) return;
    agg.provincias.add(p.province);
    agg.localidades.add(p.name);
    p.clients.forEach(name => {
      const k = contactKey(p, name, 'C');
      const done = contacted.has(k);
      agg.clientes_total++;
      if (done) agg.clientes_contactados++;
      agg.rows.push({
        Zona: (vendorLookup[vk] || {}).zone || '-',
        Vendedor: titleCase(vk),
        Provincia: titleCase(p.province),
        Departamento: p.dept || '',
        Localidad: p.name,
        Tipo: 'Cliente',
        'Nombre Comercial': name,
        Contactado: done ? 'SI' : 'NO',
        Latitud: p.lat,
        Longitud: p.lon,
      });
    });
    p.prospects.forEach(name => {
      const k = contactKey(p, name, 'P');
      const done = contacted.has(k);
      agg.prospectos_total++;
      if (done) agg.prospectos_contactados++;
      agg.rows.push({
        Zona: (vendorLookup[vk] || {}).zone || '-',
        Vendedor: titleCase(vk),
        Provincia: titleCase(p.province),
        Departamento: p.dept || '',
        Localidad: p.name,
        Tipo: 'Prospecto',
        'Nombre Comercial': name,
        Contactado: done ? 'SI' : 'NO',
        Latitud: p.lat,
        Longitud: p.lon,
      });
    });
  });

  // Summary sheet
  const sumRows = [];
  VENDORS.forEach(v => {
    const a = perVendor[v.key];
    if (!a) return;
    const total = a.clientes_total + a.prospectos_total;
    const done = a.clientes_contactados + a.prospectos_contactados;
    sumRows.push({
      Zona: a.zone,
      Vendedor: titleCase(a.vendor),
      Provincias: a.provincias.size,
      Localidades: a.localidades.size,
      'Clientes Actuales': a.clientes_total,
      'Clientes Contactados': a.clientes_contactados,
      'Clientes Pendientes': a.clientes_total - a.clientes_contactados,
      Prospectos: a.prospectos_total,
      'Prospectos Contactados': a.prospectos_contactados,
      'Prospectos Pendientes': a.prospectos_total - a.prospectos_contactados,
      'Total Puntos': total,
      'Total Contactados': done,
      '% Avance': total ? Math.round((done / total) * 100) + '%' : '-',
    });
  });
  if (perVendor['SIN ASIGNAR'].rows.length) {
    const a = perVendor['SIN ASIGNAR'];
    sumRows.push({
      Zona: '-', Vendedor: 'Sin Asignar',
      Provincias: a.provincias.size, Localidades: a.localidades.size,
      'Clientes Actuales': a.clientes_total, 'Clientes Contactados': a.clientes_contactados,
      'Clientes Pendientes': a.clientes_total - a.clientes_contactados,
      Prospectos: a.prospectos_total, 'Prospectos Contactados': a.prospectos_contactados,
      'Prospectos Pendientes': a.prospectos_total - a.prospectos_contactados,
      'Total Puntos': a.clientes_total + a.prospectos_total,
      'Total Contactados': a.clientes_contactados + a.prospectos_contactados,
      '% Avance': '-',
    });
  }
  const sumWs = XLSX.utils.json_to_sheet(sumRows);
  sumWs['!cols'] = [{wch:6},{wch:24},{wch:11},{wch:12},{wch:14},{wch:14},{wch:14},{wch:11},{wch:14},{wch:14},{wch:12},{wch:14},{wch:10}];
  XLSX.utils.book_append_sheet(wb, sumWs, 'Resumen');

  // Per-vendor detail sheets
  Object.values(perVendor).forEach(a => {
    if (!a.rows.length) return;
    a.rows.sort((x, y) => {
      if (x.Provincia !== y.Provincia) return x.Provincia.localeCompare(y.Provincia);
      if (x.Localidad !== y.Localidad) return x.Localidad.localeCompare(y.Localidad);
      if (x.Tipo !== y.Tipo) return x.Tipo.localeCompare(y.Tipo);
      return x['Nombre Comercial'].localeCompare(y['Nombre Comercial']);
    });
    const ws = XLSX.utils.json_to_sheet(a.rows);
    ws['!cols'] = [{wch:6},{wch:22},{wch:18},{wch:18},{wch:22},{wch:11},{wch:34},{wch:12},{wch:10},{wch:10}];
    let sheetName = (a.zone + ' ' + a.vendor).substring(0, 28);
    sheetName = sheetName.replace(/[\\/*?\[\]:]/g, '');
    XLSX.utils.book_append_sheet(wb, ws, sheetName);
  });

  // Master detail sheet (all rows)
  const allRows = [];
  Object.values(perVendor).forEach(a => allRows.push(...a.rows));
  if (allRows.length) {
    const wsAll = XLSX.utils.json_to_sheet(allRows);
    wsAll['!cols'] = [{wch:6},{wch:22},{wch:18},{wch:18},{wch:22},{wch:11},{wch:34},{wch:12},{wch:10},{wch:10}];
    XLSX.utils.book_append_sheet(wb, wsAll, 'Todos');
  }

  // Pedidos sheet
  const pedidoRows = [];
  Object.entries(orders).forEach(([k, lines]) => {
    if (!lines || !lines.length) return;
    const [tipo, prov, locName, clientName] = k.split('|');
    const pt = POINTS.find(p => p.province === prov && p.name === locName);
    const vendor = pt ? pt.vendor : '';
    const vm = vendorLookup[vendor];
    lines.forEach(l => {
      pedidoRows.push({
        Zona: vm ? vm.zone : '-',
        Vendedor: titleCase(vendor || ''),
        Provincia: titleCase(prov),
        Localidad: locName,
        Tipo: tipo === 'C' ? 'Cliente' : 'Prospecto',
        Cliente: clientName,
        Codigo: l.code,
        Producto: l.desc,
        Categoria: l.cat || '',
        Familia: l.fam || '',
        Subfamilia: l.sub || '',
        Cantidad: parseFloat(l.qty) || 0,
      });
    });
  });
  if (pedidoRows.length) {
    pedidoRows.sort((a, b) => {
      if (a.Vendedor !== b.Vendedor) return a.Vendedor.localeCompare(b.Vendedor);
      if (a.Cliente !== b.Cliente) return a.Cliente.localeCompare(b.Cliente);
      return a.Codigo.localeCompare(b.Codigo);
    });
    const wsP = XLSX.utils.json_to_sheet(pedidoRows);
    wsP['!cols'] = [{wch:6},{wch:22},{wch:18},{wch:22},{wch:11},{wch:30},{wch:14},{wch:38},{wch:14},{wch:18},{wch:18},{wch:10}];
    XLSX.utils.book_append_sheet(wb, wsP, 'Pedidos Pendientes');
  }

  // Confirmados sheet
  const confRows = [];
  Object.entries(confirmed).forEach(([k, list]) => {
    if (!list || !list.length) return;
    const [tipo, prov, locName, clientName] = k.split('|');
    const pt = POINTS.find(p => p.province === prov && p.name === locName);
    const vendor = pt ? pt.vendor : '';
    const vm = vendorLookup[vendor];
    list.forEach(c => {
      (c.lines || []).forEach(l => {
        confRows.push({
          Zona: vm ? vm.zone : '-',
          Vendedor: titleCase(vendor || ''),
          Provincia: titleCase(prov),
          Localidad: locName,
          Tipo: tipo === 'C' ? 'Cliente' : 'Prospecto',
          Cliente: clientName,
          Mes: c.month,
          'Fecha Confirmacion': c.confirmedAt ? new Date(c.confirmedAt).toLocaleString() : '',
          Codigo: l.code,
          Producto: l.desc,
          Categoria: l.cat || '',
          Familia: l.fam || '',
          Subfamilia: l.sub || '',
          Cantidad: parseFloat(l.qty) || 0,
        });
      });
    });
  });
  if (confRows.length) {
    confRows.sort((a, b) => {
      if (a.Mes !== b.Mes) return a.Mes.localeCompare(b.Mes);
      if (a.Vendedor !== b.Vendedor) return a.Vendedor.localeCompare(b.Vendedor);
      if (a.Cliente !== b.Cliente) return a.Cliente.localeCompare(b.Cliente);
      return a.Codigo.localeCompare(b.Codigo);
    });
    const wsC = XLSX.utils.json_to_sheet(confRows);
    wsC['!cols'] = [{wch:6},{wch:22},{wch:18},{wch:22},{wch:11},{wch:30},{wch:18},{wch:20},{wch:14},{wch:38},{wch:14},{wch:18},{wch:18},{wch:10}];
    XLSX.utils.book_append_sheet(wb, wsC, 'Pedidos Confirmados');
  }

  const today = new Date().toISOString().slice(0, 10);
  XLSX.writeFile(wb, 'Shimano_Zonas_Contactos_' + today + '.xlsx');
};

window.setTab = function(name){
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.toggle('active', b.dataset.tab === name));
  document.getElementById('pane-locs').style.display = name === 'locs' ? '' : 'none';
  document.getElementById('pane-clients').style.display = name === 'clients' ? '' : 'none';
  document.getElementById('pane-pedidos').style.display = name === 'pedidos' ? '' : 'none';
  if (name === 'pedidos') renderPedidosTab();
};

// === Pedidos tab: lista de clientes contactados con su contador de pedido ===
function getContactedClientsFromFilter(pts){
  // Devuelve [{clientName, tipo, point}] solo con los contactados
  const items = [];
  pts.forEach(p => {
    p.clients.forEach(n => {
      if (contacted.has(contactKey(p, n, 'C'))) items.push({name:n, tipo:'C', point:p});
    });
    p.prospects.forEach(n => {
      if (contacted.has(contactKey(p, n, 'P'))) items.push({name:n, tipo:'P', point:p});
    });
  });
  return items;
}

let pedidoView = 'crear';
window.setPedidoView = function(v){
  pedidoView = v;
  document.getElementById('pt-crear').classList.toggle('active', v === 'crear');
  document.getElementById('pt-pend').classList.toggle('active', v === 'pendientes');
  document.getElementById('pt-conf').classList.toggle('active', v === 'confirmados');
  renderPedidosTab();
};

function renderPedidosTab(){
  if (pedidoView === 'confirmados') return renderConfirmadosList();
  if (pedidoView === 'crear') return renderCrearList();
  return renderPendientesList();
}

function renderCrearList(){
  const pts = filteredPoints();
  const items = getContactedClientsFromFilter(pts);
  items.sort((a, b) => a.name.localeCompare(b.name));
  let html = '';
  items.forEach(it => {
    const k = orderKey(it.name, it.point.province, it.point.name, it.tipo);
    const hasOrder = (orders[k] || []).length > 0;
    const tipoLabel = it.tipo === 'C' ? 'Cliente' : 'Prospecto';
    const tipoBg = it.tipo === 'C' ? '#00A9E0' : '#f59e0b';
    html += '<div class="pedido-client-card" ondblclick="askCreatePedido(\'' + escapeAttr(it.name) + '\',\'' + it.tipo + '\',\'' + escapeAttr(it.point.province) + '\',\'' + escapeAttr(it.point.name) + '\')" title="Doble click para crear pedido">';
    html += '<div style="flex:1;min-width:0"><div class="ped-name">' + escapeHtml(it.name) + '</div>';
    html += '<div class="ped-meta">' + escapeHtml(it.point.name) + ' / ' + escapeHtml(titleCase(it.point.province)) + '</div></div>';
    html += '<div style="text-align:right;display:flex;flex-direction:column;gap:3px;align-items:flex-end">';
    html += '<div style="background:' + tipoBg + ';color:#fff;font-size:9px;font-weight:800;padding:2px 8px;border-radius:10px;text-transform:uppercase;letter-spacing:.4px">' + tipoLabel + '</div>';
    if (hasOrder) html += '<div style="font-size:9px;color:#166534;font-weight:700">En curso</div>';
    html += '</div></div>';
  });
  if (!items.length) html = '<div style="color:#94a3b8;font-size:12px;padding:14px;text-align:center;background:#f8fafc;border-radius:6px;line-height:1.5">Sin clientes contactados en la seleccion actual.<br><span style="font-size:10px">Marcalos primero en el tab <b>Clientes</b>.</span></div>';
  document.getElementById('pedidos-content').innerHTML = html;
  document.getElementById('tab-ped-count').textContent = items.length;
  document.getElementById('pedidos-summary').textContent = items.length + ' cliente(s) contactado(s). Doble click para crear pedido.';
}

window.askCreatePedido = function(name, tipo, province, locName){
  if (!confirm('Queres crear un pedido para "' + name + '"?')) return;
  openPedidoModal(name, tipo, province, locName);
};

function renderPendientesList(){
  const pts = filteredPoints();
  const items = getContactedClientsFromFilter(pts).filter(it => {
    const k = orderKey(it.name, it.point.province, it.point.name, it.tipo);
    return (orders[k] || []).length > 0;
  });
  items.sort((a, b) => a.name.localeCompare(b.name));
  let html = '';
  items.forEach(it => {
    const k = orderKey(it.name, it.point.province, it.point.name, it.tipo);
    const order = orders[k] || [];
    const totalU = order.reduce((s, l) => s + (parseFloat(l.qty) || 0), 0);
    html += '<div class="pedido-client-card" onclick="openPedidoModal(\'' + escapeAttr(it.name) + '\',\'' + it.tipo + '\',\'' + escapeAttr(it.point.province) + '\',\'' + escapeAttr(it.point.name) + '\')" title="Click para editar el pedido">';
    html += '<div style="flex:1;min-width:0"><div class="ped-name">' + escapeHtml(it.name) + '</div>';
    html += '<div class="ped-meta">' + escapeHtml(it.point.name) + ' / ' + escapeHtml(titleCase(it.point.province)) + '</div></div>';
    html += '<div style="text-align:right"><div class="ped-badge">' + order.length + ' items</div>';
    html += '<div style="font-size:9px;color:#64748b;margin-top:2px">' + totalU + ' unid.</div>';
    html += '</div></div>';
  });
  if (!items.length) html = '<div style="color:#94a3b8;font-size:12px;padding:14px;text-align:center;background:#f8fafc;border-radius:6px;line-height:1.5">Sin pedidos en curso.<br><span style="font-size:10px">Andate al sub-tab <b>Crear</b> para iniciar uno.</span></div>';
  document.getElementById('pedidos-content').innerHTML = html;
  document.getElementById('tab-ped-count').textContent = items.length;
  document.getElementById('pedidos-summary').textContent = items.length + ' pedido(s) en curso (sin confirmar)';
}

function renderConfirmadosList(){
  const pts = filteredPoints();
  const visibleKeys = new Set();
  pts.forEach(p => {
    p.clients.forEach(n => visibleKeys.add(orderKey(n, p.province, p.name, 'C')));
    p.prospects.forEach(n => visibleKeys.add(orderKey(n, p.province, p.name, 'P')));
  });
  const items = [];
  Object.entries(confirmed).forEach(([k, list]) => {
    if (!visibleKeys.has(k)) return;
    const [tipo, prov, locName, clientName] = k.split('|');
    (list || []).forEach((c, idx) => {
      items.push({key: k, idx: idx, tipo, prov, locName, clientName, conf: c});
    });
  });
  items.sort((a, b) => (b.conf.confirmedAt || '').localeCompare(a.conf.confirmedAt || ''));
  let html = '';
  items.forEach(it => {
    const lineCount = (it.conf.lines || []).length;
    const units = (it.conf.lines || []).reduce((s, l) => s + (parseFloat(l.qty) || 0), 0);
    html += '<div class="confirmed-card" onclick="viewConfirmed(\'' + escapeAttr(it.key) + '\',' + it.idx + ')">';
    html += '<div class="cc-head"><div class="cc-name">' + escapeHtml(it.clientName) + '</div><div class="cc-month">' + escapeHtml(it.conf.month) + '</div></div>';
    html += '<div class="cc-meta">' + escapeHtml(it.locName) + ' / ' + escapeHtml(titleCase(it.prov)) + '</div>';
    html += '<div class="cc-summary">' + lineCount + ' productos &middot; ' + units + ' unidades &middot; ' + (it.conf.confirmedAt ? new Date(it.conf.confirmedAt).toLocaleDateString() : '') + '</div>';
    html += '</div>';
  });
  if (!items.length) html = '<div style="color:#94a3b8;font-size:12px;padding:14px;text-align:center;background:#f8fafc;border-radius:6px;line-height:1.5">No hay pedidos confirmados en la seleccion actual.</div>';
  document.getElementById('pedidos-content').innerHTML = html;
  document.getElementById('pedidos-summary').textContent = items.length + ' pedido(s) confirmado(s)';
}

window.viewConfirmed = function(key, idx){
  const list = confirmed[key] || [];
  const c = list[idx];
  if (!c) return;
  const [tipo, prov, locName, clientName] = key.split('|');
  // open pedido modal in read-only mode
  currentOrderKey = '__readonly__';
  currentOrderClient = {name: clientName, tipo: tipo, province: prov, locName: locName};
  document.getElementById('pm-name').textContent = clientName + ' - ' + c.month;
  const tipoLbl = tipo === 'C' ? 'Cliente actual' : 'Prospecto';
  document.getElementById('pm-meta').innerHTML = '<span class="badge tipo">' + tipoLbl + '</span>'
    + '<span class="badge prov">' + escapeHtml(titleCase(prov)) + '</span>'
    + '<span>' + escapeHtml(locName) + '</span>'
    + '<span style="background:#166534;color:#fff;padding:2px 8px;border-radius:10px;font-size:10px;font-weight:800">CONFIRMADO</span>';
  // Show lines as read-only
  let html = '';
  let totalU = 0;
  (c.lines || []).forEach(l => {
    const q = parseFloat(l.qty) || 0;
    totalU += q;
    html += '<div class="ped-line" style="grid-template-columns:75px 1fr 56px">';
    html += '<div class="pcode">' + escapeHtml(l.code) + '</div>';
    html += '<div class="pname">' + escapeHtml(l.desc) + '</div>';
    html += '<div style="text-align:right;font-weight:800;color:#166534">' + q + '</div>';
    html += '</div>';
  });
  document.getElementById('pm-lines').innerHTML = html || '<div class="no-data">Pedido vacio</div>';
  document.getElementById('pm-line-count').textContent = (c.lines || []).length + ' productos';
  document.getElementById('pm-units').textContent = totalU + ' unidades';
  document.getElementById('pm-products').innerHTML = '<div class="no-data">Pedido CONFIRMADO (solo lectura).<br>Confirmado el ' + (c.confirmedAt ? new Date(c.confirmedAt).toLocaleString() : '-') + '</div>';
  document.getElementById('pm-confirm').style.display = 'none';
  document.getElementById('pm-cancel').style.display = 'none';
  document.getElementById('pm-saved-tag').textContent = 'Solo lectura';
  document.getElementById('pm-suggest-box').classList.add('hidden');
  document.getElementById('pedido-modal').classList.add('open');
};

// === Modal de pedido: filtros + picker + lineas ===
let currentOrderKey = null;
let currentOrderClient = null;

function openPedidoModal(clientName, tipo, province, locName){
  currentOrderKey = orderKey(clientName, province, locName, tipo);
  currentOrderClient = {name: clientName, tipo: tipo, province: province, locName: locName};
  document.getElementById('pm-name').textContent = clientName;
  const tipoLbl = tipo === 'C' ? 'Cliente actual' : 'Prospecto';
  document.getElementById('pm-meta').innerHTML = '<span class="badge tipo">' + tipoLbl + '</span>'
    + '<span class="badge prov">' + escapeHtml(titleCase(province)) + '</span>'
    + '<span>' + escapeHtml(locName) + '</span>';
  document.getElementById('pm-search').value = '';
  populateProductFilters();
  renderProductPicker();
  renderOrderLines();
  renderSuggestions();
  document.getElementById('pm-saved-tag').textContent = 'Sin cambios';
  document.getElementById('pm-confirm').style.display = '';
  document.getElementById('pm-cancel').style.display = '';
  document.getElementById('pedido-modal').classList.add('open');
}

window.cancelPedido = function(){
  if (!currentOrderKey || currentOrderKey === '__readonly__') return;
  const ord = orders[currentOrderKey] || [];
  if (!ord.length) {
    closePedidoModal();
    return;
  }
  if (!confirm('Eliminar TODO el pedido de "' + currentOrderClient.name + '"? Esta accion no se puede deshacer.')) return;
  delete orders[currentOrderKey];
  saveOrders();
  closePedidoModal();
};

window.openConfirmDialog = function(){
  if (!currentOrderKey || currentOrderKey === '__readonly__') return;
  const ord = orders[currentOrderKey] || [];
  if (!ord.length) { alert('El pedido esta vacio. Agregue productos antes de confirmar.'); return; }
  // populate month/year selects
  const now = new Date();
  const mesSel = document.getElementById('cd-mes');
  mesSel.innerHTML = MESES.map((m, i) => '<option value="' + i + '">' + m + '</option>').join('');
  mesSel.value = now.getMonth();
  const anioSel = document.getElementById('cd-anio');
  const year = now.getFullYear();
  let yopts = '';
  for (let y = year - 1; y <= year + 2; y++) yopts += '<option value="' + y + '">' + y + '</option>';
  anioSel.innerHTML = yopts;
  anioSel.value = year;
  document.getElementById('cd-title').textContent = 'Confirmar pedido';
  document.getElementById('cd-text').innerHTML = 'Desea confirmar el pedido de <b>' + escapeHtml(currentOrderClient.name) + '</b>?<br><span style="font-size:11px;color:#64748b">Seleccione el mes y a&ntilde;o:</span>';
  document.getElementById('confirm-dialog').classList.add('open');
};

window.closeConfirmDialog = function(){
  document.getElementById('confirm-dialog').classList.remove('open');
};

window.doConfirmPedido = function(){
  if (!currentOrderKey || currentOrderKey === '__readonly__') return;
  const ord = orders[currentOrderKey] || [];
  if (!ord.length) return;
  const mesIdx = parseInt(document.getElementById('cd-mes').value);
  const anio = document.getElementById('cd-anio').value;
  const monthLabel = MESES[mesIdx] + ' ' + anio;
  if (!confirm('Confirmar pedido de "' + currentOrderClient.name + '" para ' + monthLabel + '?')) return;
  if (!confirmed[currentOrderKey]) confirmed[currentOrderKey] = [];
  confirmed[currentOrderKey].push({
    month: monthLabel,
    monthIdx: mesIdx,
    year: parseInt(anio),
    confirmedAt: new Date().toISOString(),
    lines: ord.map(l => ({code: l.code, desc: l.desc, cat: l.cat, fam: l.fam, sub: l.sub, qty: parseFloat(l.qty) || 0})),
  });
  saveConfirmed();
  delete orders[currentOrderKey];
  saveOrders();
  closeConfirmDialog();
  closePedidoModal();
  setPedidoView('confirmados');
};
window.openPedidoModal = openPedidoModal;

function closePedidoModal(){
  document.getElementById('pedido-modal').classList.remove('open');
  currentOrderKey = null;
  currentOrderClient = null;
  renderPedidosTab();
}
window.closePedidoModal = closePedidoModal;

function populateProductFilters(){
  const cats = [...new Set(PRODUCTS.map(p => p.cat).filter(Boolean))].sort();
  const fams = [...new Set(PRODUCTS.map(p => p.fam).filter(Boolean))].sort();
  const subs = [...new Set(PRODUCTS.map(p => p.sub).filter(Boolean))].sort();
  document.getElementById('pm-cat').innerHTML = '<option value="ALL">Categoria: Todas</option>' + cats.map(c => '<option value="' + escapeHtml(c) + '">' + escapeHtml(c) + '</option>').join('');
  document.getElementById('pm-fam').innerHTML = '<option value="ALL">Familia: Todas</option>' + fams.map(c => '<option value="' + escapeHtml(c) + '">' + escapeHtml(c) + '</option>').join('');
  document.getElementById('pm-sub').innerHTML = '<option value="ALL">Subfamilia: Todas</option>' + subs.map(c => '<option value="' + escapeHtml(c) + '">' + escapeHtml(c) + '</option>').join('');
}

function renderProductPicker(){
  const cat = document.getElementById('pm-cat').value;
  const fam = document.getElementById('pm-fam').value;
  const sub = document.getElementById('pm-sub').value;
  const q = (document.getElementById('pm-search').value || '').toLowerCase().trim();
  const inOrder = new Set((orders[currentOrderKey] || []).map(l => l.code));
  const filt = PRODUCTS.filter(p => {
    if (cat !== 'ALL' && p.cat !== cat) return false;
    if (fam !== 'ALL' && p.fam !== fam) return false;
    if (sub !== 'ALL' && p.sub !== sub) return false;
    if (q && !(p.code.toLowerCase().includes(q) || p.desc.toLowerCase().includes(q))) return false;
    return true;
  }).slice(0, 200);
  let html = '';
  filt.forEach(p => {
    const cls = inOrder.has(p.code) ? 'prod-row in-order' : 'prod-row';
    html += '<div class="' + cls + '" onclick="addToOrder(\'' + escapeAttr(p.code) + '\')">';
    html += '<div class="code">' + escapeHtml(p.code) + '</div>';
    html += '<div><div class="pdesc">' + escapeHtml(p.desc) + '</div>';
    html += '<div class="pcat"><span>' + escapeHtml(p.cat) + '</span><span>' + escapeHtml(p.fam) + '</span><span>' + escapeHtml(p.sub) + '</span></div></div>';
    html += '<button class="add-btn" onclick="event.stopPropagation();addToOrder(\'' + escapeAttr(p.code) + '\')">+</button></div>';
  });
  if (!filt.length) html = '<div class="no-data">Sin productos para estos filtros.</div>';
  document.getElementById('pm-products').innerHTML = html;
}
window.renderProductPicker = renderProductPicker;

function addToOrder(code){
  if (!currentOrderKey || currentOrderKey === '__readonly__') return;
  if (!orders[currentOrderKey]) orders[currentOrderKey] = [];
  const ord = orders[currentOrderKey];
  const existing = ord.find(l => l.code === code);
  if (existing) {
    existing.qty = (parseFloat(existing.qty) || 0) + 1;
  } else {
    const prod = PRODUCTS.find(p => p.code === code);
    if (!prod) return;
    ord.push({code: prod.code, desc: prod.desc, cat: prod.cat, fam: prod.fam, sub: prod.sub, qty: 1});
  }
  saveOrders();
  flashSaved();
  renderOrderLines();
  renderProductPicker();
  renderSuggestions();
}
window.addToOrder = addToOrder;

function setOrderQty(code, qty){
  if (!currentOrderKey) return;
  const ord = orders[currentOrderKey] || [];
  const line = ord.find(l => l.code === code);
  if (!line) return;
  const q = parseFloat(qty);
  if (isNaN(q) || q <= 0) {
    orders[currentOrderKey] = ord.filter(l => l.code !== code);
  } else {
    line.qty = q;
  }
  saveOrders();
  flashSaved();
  renderOrderLines();
}
window.setOrderQty = setOrderQty;

function removeFromOrder(code){
  if (!currentOrderKey || currentOrderKey === '__readonly__') return;
  orders[currentOrderKey] = (orders[currentOrderKey] || []).filter(l => l.code !== code);
  saveOrders();
  flashSaved();
  renderOrderLines();
  renderProductPicker();
  renderSuggestions();
}
window.removeFromOrder = removeFromOrder;

function normTitle(s){
  return String(s || '').normalize('NFD').replace(/[̀-ͯ]/g, '').toUpperCase().replace(/[^A-Z0-9]/g, '');
}

// Indice: SKU code (normalizado) -> producto
const SKU_INDEX = (() => {
  const map = {};
  PRODUCTS.forEach(p => {
    const k = normTitle(p.code);
    if (k.length >= 3) map[k] = p;
  });
  return map;
})();

// Indice por familia/subfamilia: { token: [products] } para matcheo por descripcion
const SKU_TOKENS = (() => {
  const map = {};
  PRODUCTS.forEach(p => {
    const tokens = new Set();
    [p.sub, p.fam].forEach(s => {
      if (!s) return;
      const norm = normTitle(s);
      if (norm.length >= 3) tokens.add(norm);
    });
    tokens.forEach(t => {
      if (!map[t]) map[t] = [];
      map[t].push(p);
    });
  });
  return map;
})();

function matchSkuFromTitle(meliTitle){
  const t = normTitle(meliTitle);
  // Paso 1: SKU code substring (mas especifico)
  let best = null;
  let bestLen = 0;
  for (const k in SKU_INDEX) {
    if (t.indexOf(k) !== -1 && k.length > bestLen) {
      best = SKU_INDEX[k];
      bestLen = k.length;
    }
  }
  if (best) return best;
  // Paso 2: familia/subfamilia substring (mas laxo)
  for (const tok in SKU_TOKENS) {
    if (t.indexOf(tok) !== -1) {
      return SKU_TOKENS[tok][0];
    }
  }
  return null;
}

function renderSuggestions(){
  if (!currentOrderClient || currentOrderKey === '__readonly__') return;
  const box = document.getElementById('pm-suggest-box');
  const listEl = document.getElementById('pm-suggest');
  const infoEl = document.getElementById('pm-suggest-info');
  const province = currentOrderClient.province;

  // Agregar productos de pedidos CONFIRMADOS de otras casas de pesca de la misma provincia
  const agg = {}; // code -> {qty, shops:Set, sampleLine}
  const peerShops = new Set();
  Object.entries(confirmed).forEach(([k, list]) => {
    if (!list || !list.length) return;
    const parts = k.split('|');
    const prov = parts[1];
    const clientName = parts[3];
    if (prov !== province) return;
    if (clientName === currentOrderClient.name) return; // excluir cliente actual
    peerShops.add(clientName);
    list.forEach(c => {
      (c.lines || []).forEach(l => {
        if (!l.code) return;
        if (!agg[l.code]) agg[l.code] = {qty: 0, shops: new Set(), line: l};
        agg[l.code].qty += parseFloat(l.qty) || 0;
        agg[l.code].shops.add(clientName);
      });
    });
  });

  // SKUs ya en el pedido actual
  const inOrder = new Set((orders[currentOrderKey] || []).map(l => l.code));

  // Ordenar por # casas (popularidad) y luego por unidades
  const suggestions = Object.entries(agg)
    .filter(([code]) => !inOrder.has(code))
    .map(([code, d]) => ({code, qty: d.qty, shops: d.shops.size, shopList: [...d.shops].sort(), line: d.line}))
    .sort((a, b) => (b.shops - a.shops) || (b.qty - a.qty))
    .slice(0, 8);

  box.classList.remove('hidden');
  infoEl.textContent = peerShops.size + ' casa(s) confirmadas en ' + titleCase(province);

  if (!suggestions.length) {
    let msg;
    if (peerShops.size === 0) {
      msg = 'Aun no hay pedidos confirmados de otras casas de pesca en <b>' + escapeHtml(titleCase(province)) + '</b>.<br><span style="font-size:9px">Las sugerencias se construyen a partir de los pedidos confirmados en el sistema.</span>';
    } else {
      msg = 'Las casas de pesca de <b>' + escapeHtml(titleCase(province)) + '</b> ya tienen sus productos cubiertos por el pedido actual.';
    }
    listEl.innerHTML = '<div class="suggest-empty-msg">' + msg + '</div>';
    return;
  }

  let html = '';
  suggestions.forEach(s => {
    const masterProd = PRODUCTS.find(p => p.code === s.code);
    const desc = (masterProd && masterProd.desc) || s.line.desc || s.code;
    // Mostrar nombres de las casas: hasta 3, sino "X, Y y N mas"
    let shopsText;
    if (s.shopList.length <= 3) {
      shopsText = s.shopList.join(', ');
    } else {
      shopsText = s.shopList.slice(0, 2).join(', ') + ' y ' + (s.shopList.length - 2) + ' mas';
    }
    const shopsTitle = s.shopList.join(' | ');
    html += '<div class="suggest-row" onclick="addToOrder(\'' + escapeAttr(s.code) + '\')">';
    html += '<div class="code">' + escapeHtml(s.code) + '</div>';
    html += '<div><div class="sname">' + escapeHtml(desc) + '</div>';
    html += '<div class="sinfo" title="' + escapeHtml(shopsTitle) + '">Pidieron: ' + escapeHtml(shopsText) + ' &middot; ' + s.qty + ' unid.</div></div>';
    html += '<button class="add-sg" onclick="event.stopPropagation();addToOrder(\'' + escapeAttr(s.code) + '\')" title="Agregar al pedido">+</button>';
    html += '</div>';
  });
  listEl.innerHTML = html;
}

function renderOrderLines(){
  const ord = orders[currentOrderKey] || [];
  let totalU = 0;
  let html = '';
  ord.forEach(l => {
    const q = parseFloat(l.qty) || 0;
    totalU += q;
    html += '<div class="ped-line">';
    html += '<div class="pcode">' + escapeHtml(l.code) + '</div>';
    html += '<div class="pname">' + escapeHtml(l.desc) + '</div>';
    html += '<input type="number" class="qty" min="0" step="1" value="' + q + '" onchange="setOrderQty(\'' + escapeAttr(l.code) + '\', this.value)"/>';
    html += '<button class="rm-btn" onclick="removeFromOrder(\'' + escapeAttr(l.code) + '\')" title="Quitar">&times;</button>';
    html += '</div>';
  });
  if (!ord.length) html = '<div class="no-data">Sin productos cargados. Agregue desde la izquierda.</div>';
  document.getElementById('pm-lines').innerHTML = html;
  document.getElementById('pm-line-count').textContent = ord.length + ' producto(s) en pedido';
  document.getElementById('pm-units').textContent = totalU + ' unidades';
}

let savedTimer = null;
function flashSaved(){
  const el = document.getElementById('pm-saved-tag');
  if (!el) return;
  el.textContent = 'Guardado ' + new Date().toLocaleTimeString();
  clearTimeout(savedTimer);
  savedTimer = setTimeout(() => { el.textContent = 'Sin cambios'; }, 2500);
}

function updateSidebar(pts){
  const title = document.getElementById('sidebar-title');
  if (currentLocality !== 'ALL') title.textContent = 'Localidad: ' + currentLocality;
  else if (currentProvince !== 'ALL') title.textContent = 'Localidades en ' + titleCase(currentProvince);
  else if (currentVendor !== 'ALL') {
    const v = vendorLookup[currentVendor];
    title.textContent = v ? v.zone + ' - ' + titleCase(currentVendor) : 'Localidades';
  } else title.textContent = 'Localidades (vista nacional)';

  const sorted = [...pts].sort((a, b) => (b.clients.length + b.prospects.length) - (a.clients.length + a.prospects.length));
  let html = '';
  for (const p of sorted) {
    const c = p.clients.length, pp = p.prospects.length, total = c + pp;
    const vm = vendorLookup[p.vendor];
    const accent = vm ? vm.color : '#00A9E0';
    html += '<div class="loc-card" style="border-left-color:' + accent + '" onclick="zoomTo(' + p.lat + ',' + p.lon + ')">';
    html += '<div class="loc-card-title">' + escapeHtml(p.name) + '<span class="badge" style="background:' + accent + '">' + total + '</span></div>';
    html += '<div class="loc-card-meta"><span class="c">' + c + ' clientes</span> &nbsp;-&nbsp; <span class="p">' + pp + ' prospectos</span></div>';
    html += '<div class="loc-card-meta" style="font-size:10px;margin-top:2px;color:#94a3b8">' + escapeHtml(p.province) + (p.dept ? ' / ' + escapeHtml(p.dept) : '') + '</div>';
    html += '</div>';
  }
  if (!sorted.length) html = '<div style="color:#94a3b8;font-size:12px;padding:8px">Sin localidades con datos para esta combinacion.</div>';
  document.getElementById('sidebar-content').innerHTML = html;
  document.getElementById('tab-loc-count').textContent = sorted.length;
  renderClients(pts);
  renderPedidosTab();
}

window.zoomTo = function(lat, lon){
  map.flyTo([lat, lon], 11, {duration: .7});
};

function setFilter(f){
  currentFilter = f;
  document.querySelectorAll('.filter-btn').forEach(b => b.classList.toggle('active', b.dataset.filter === f));
  drawMarkers();
}
window.setFilter = setFilter;

function getFilteredBounds(){
  // Builds bounds from dept polygons that match current filters; falls back to point coords.
  const coords = [];
  DEPT_GEO.features.forEach(f => {
    if (currentVendor !== 'ALL' && f.properties.vendor !== currentVendor) return;
    if (currentProvince !== 'ALL' && f.properties.province !== currentProvince) return;
    if (!deptMatchesLocality(f)) return;
    const flatten = (arr) => {
      if (typeof arr[0] === 'number') { coords.push([arr[1], arr[0]]); return; }
      arr.forEach(flatten);
    };
    flatten(f.geometry.coordinates);
  });
  if (!coords.length) {
    filteredPoints().forEach(p => coords.push([p.lat, p.lon]));
  }
  return coords.length ? L.latLngBounds(coords) : null;
}

function recenter(){
  if (currentVendor === 'ALL' && currentProvince === 'ALL' && currentLocality === 'ALL') {
    map.setView([-38, -64], 4);
    return;
  }
  if (currentLocality !== 'ALL') {
    const p = POINTS.find(pt => pt.name === currentLocality && (currentProvince === 'ALL' || pt.province === currentProvince));
    if (p) { map.flyTo([p.lat, p.lon], 12, {duration: .7}); return; }
  }
  const b = getFilteredBounds();
  if (b) map.fitBounds(b, {padding: [40, 40], maxZoom: currentProvince !== 'ALL' ? 10 : 9});
}

function showClickDebug(msg){
  const el = document.getElementById('click-debug');
  if (!el) return;
  el.textContent = msg;
  el.style.display = 'block';
  clearTimeout(window.__clickDebugTimer);
  window.__clickDebugTimer = setTimeout(() => { el.style.display = 'none'; }, 3000);
}

function selectFromMap(opts){
  // opts: {vendor, province, locality} - any subset; missing = reset to ALL
  if ('vendor' in opts) {
    currentVendor = opts.vendor || 'ALL';
    document.getElementById('zone-select').value = currentVendor;
  }
  if ('province' in opts) currentProvince = opts.province || 'ALL';
  else currentProvince = 'ALL';
  if ('locality' in opts) currentLocality = opts.locality || 'ALL';
  else currentLocality = 'ALL';
  populateProvinces();
  document.getElementById('prov-select').value = currentProvince;
  populateLocalities();
  document.getElementById('loc-select').value = currentLocality;
  restyleZoneLayers();
  recenter();
  drawMarkers();
}
window.selectFromMap = selectFromMap;

document.getElementById('zone-select').addEventListener('change', (e) => {
  currentVendor = e.target.value;
  currentProvince = 'ALL';
  currentLocality = 'ALL';
  populateProvinces();
  populateLocalities();
  restyleZoneLayers();
  recenter();
  drawMarkers();
});

document.getElementById('prov-select').addEventListener('change', (e) => {
  currentProvince = e.target.value;
  currentLocality = 'ALL';
  populateLocalities();
  restyleZoneLayers();
  recenter();
  drawMarkers();
});

document.getElementById('loc-select').addEventListener('change', (e) => {
  currentLocality = e.target.value;
  restyleZoneLayers();
  recenter();
  drawMarkers();
});

map.on('zoomend', drawMarkers);

populateProvinces();
populateLocalities();
drawMarkers();
</script>
</body>
</html>
'''

html = html_template.replace('__VENDOR_OPTIONS__', vendor_options_html)
html = html.replace('__POINTS_JSON__', points_json)
html = html.replace('__VENDORS_JSON__', vendors_json)
html = html.replace('__VENDOR_AGG_JSON__', vendor_agg_json)
html = html.replace('__PROV_AGG_JSON__', prov_agg_json)
html = html.replace('__DEPT_GEO_JSON__', dept_geo_json)
html = html.replace('__PROV_GEO_JSON__', prov_geo_json)
html = html.replace('__CLIENT_SALES_JSON__', client_sales_json)
html = html.replace('__PROV_TOP_PRODUCTS_JSON__', prov_top_products_json)
html = html.replace('__PRODUCTS_JSON__', products_json)

with open(OUT, 'w', encoding='utf-8') as f:
    f.write(html)

print(f'\nWROTE: {OUT}  ({len(html)} chars)')
